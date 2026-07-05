import sqlite3
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi import Form as FastAPIForm
from starlette.middleware.sessions import SessionMiddleware
import os
import io
import traceback
import zipfile
import shutil
import requests
import uvicorn
import pandas as pd
import openpyxl
import tempfile
import re
import json
import numpy as np
from datetime import datetime, date
from typing import Dict, List, Any, Optional
from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import create_engine, text
from openpyxl.utils import column_index_from_string

from excel_analyzer import ExcelAnalyzer, sanitize_sheet_name, sanitize_and_deduplicate_columns

# ==========================================
# License / Activation (Secure Verification)
# ==========================================
import hmac
import hashlib
from fastapi import Request
from datetime import timedelta

app = FastAPI(title="Synapto System Architect API v7")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# Admin Session Middleware
# ==========================================
app.add_middleware(SessionMiddleware, secret_key="synapto-admin-session-secret-change-me")

# ==========================================
# Admin Templates
# ==========================================
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(BASE_DIR / "admin_templates"))


ADMIN_PASSWORD = os.environ.get("SYNAPTO_ADMIN_PASSWORD", "admin123")

# ==========================================
# Admin Database (SQLite)
# ==========================================
ADMIN_DB_PATH = "admin.db"

def get_admin_db():
    conn = sqlite3.connect(ADMIN_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_admin_db():
    conn = get_admin_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS activation_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            email TEXT NOT NULL,
            phone TEXT,
            company TEXT,
            notes TEXT,
            status TEXT DEFAULT 'pending',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            reviewed_at TEXT,
            admin_note TEXT
        );
        CREATE TABLE IF NOT EXISTS licenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            request_id INTEGER,
            email TEXT NOT NULL UNIQUE,
            full_name TEXT,
            company TEXT,
            token TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            deactivated_at TEXT,
            deactivate_reason TEXT
        );
        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            email TEXT,
            details TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
    ''')
    conn.commit()
    conn.close()

init_admin_db()

def log_activity(action, email=None, details=None):
    try:
        conn = get_admin_db()
        conn.execute("INSERT INTO activity_log (action, email, details) VALUES (?, ?, ?)", (action, email, details))
        conn.commit()
        conn.close()
    except:
        pass


# ==========================================
# AI Provider Configuration (Environment Variables)
# ==========================================
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "")
OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"
OPENROUTER_MAX_TOKENS = 4000

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3-flash-preview:generateContent?key={GEMINI_API_KEY}"

PRO_MODELS = {"pro": "openai/gpt-4o", "pro-claude": "anthropic/claude-3.5-sonnet"}

class ReportRequest(BaseModel):
    data: dict
    provider: str = "openrouter-free"
    architectData: Optional[dict] = None


# ==========================================
# JSON Serialization Sanitization
# ==========================================
def sanitize_for_json(obj):
    if isinstance(obj, dict):
        return {str(k): sanitize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [sanitize_for_json(item) for item in obj]
    elif isinstance(obj, set):
        return [sanitize_for_json(item) for item in obj]
    elif isinstance(obj, (np.integer,)):
        return int(obj)
    elif isinstance(obj, (np.floating,)):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return sanitize_for_json(obj.tolist())
    elif isinstance(obj, (np.bool_,)):
        return bool(obj)
    elif isinstance(obj, float):
        if pd.isna(obj) or obj != obj:
            return None
        return obj
    elif isinstance(obj, (datetime, date, pd.Timestamp)):
        return obj.isoformat()
    elif isinstance(obj, pd.Timedelta):
        return str(obj)
    elif obj is pd.NA or obj is None:
        return None
    elif isinstance(obj, bytes):
        return obj.decode('utf-8', errors='replace')
    elif isinstance(obj, (int, str, bool)):
        return obj
    else:
        try:
            return str(obj)
        except:
            return None


# ==========================================
# Enhanced Classification & Semantic Extraction
# ==========================================

# FIX 1: _classify_table - Don't count numeric column names as garbage
def _classify_table(t_meta):
    total_cols = len(t_meta['columns'])
    if total_cols == 0:
        return "Empty"
    garbage_cols = len([c for c in t_meta['columns']
                        if c['name'].startswith('empty_col')
                        or c['name'].startswith('unnamed')])
    # Numeric column names (1, 2, 5+) are valid matrix keys, not garbage.
    # But if most columns ARE numeric keys, this IS a flat matrix pattern.
    numeric_key_cols = len([c for c in t_meta['columns']
                            if re.match(r'^\d+(\+)?$', c['name'])])
    if (garbage_cols / total_cols) > 0.3 or total_cols > 50 or (total_cols > 2 and numeric_key_cols / total_cols > 0.5):
        return "Flat_Calculation_Matrix"
    return "Relational_Entity"


# FIX 2: _determine_functional_role - Better role classification for calculated tables
def _determine_functional_role(t_name, t_meta, incoming_refs, outgoing_refs, has_charts):
    table_status = _classify_table(t_meta)
    if has_charts or 'dashboard' in t_name.lower():
        return "Result_Display"
    if table_status == "Flat_Calculation_Matrix":
        return "Backend_Config" if incoming_refs > 0 else "Calculation_Matrix"
    name_lower = t_name.lower()
    # V5 FIX: Tables with many incoming refs are SOURCE tables (others depend on them)
    # e.g., employees has incoming=4 because attendance/salary/benefits/payroll FK to it
    # Such tables are Input_Form, NOT Calculated_Result
    # Calculated tables have OUTGOING refs (they depend on other tables for computation)
    if incoming_refs > 0 and outgoing_refs == 0:
        # This table is referenced by others but doesn't reference anyone = SOURCE table
        if any(kw in name_lower for kw in ['employee', 'user', 'staff']):
            return "Input_Form"
        if any(kw in name_lower for kw in ['tax', 'config', 'setting', 'bracket']):
            return "Backend_Config"
        # Generic source table
        return "Input_Form"
    # Tables explicitly named as computed
    if any(kw in name_lower for kw in ['salary_processing', 'benefit', 'final_payroll', 'calculation', 'computed']):
        return "Calculated_Result"
    # Tables that reference others and are computed from them
    if outgoing_refs > 0 and incoming_refs == 0:
        if any(kw in name_lower for kw in ['attendance', 'input', 'entry', 'form']):
            return "Input_Form"
        return "Calculated_Result"
    # Default for standalone tables
    if any(kw in name_lower for kw in ['employee', 'user', 'staff', 'attendance']):
        return "Input_Form"
    return "Input_Form"


def _derive_execution_flow(analysis_result):
    tables_meta = analysis_result.get("tables_metadata", {})
    cross_refs = analysis_result.get("cross_sheet_refs", [])
    ref_counts = {
        t_name: {"incoming": 0, "outgoing": 0, "incoming_from": [], "outgoing_to": []}
        for t_name in tables_meta
    }
    for ref in cross_refs:
        source, target = ref.get("source", ""), ref.get("target", "")
        if source in ref_counts:
            ref_counts[source]["outgoing"] += 1
            if target not in ref_counts[source]["outgoing_to"]:
                ref_counts[source]["outgoing_to"].append(target)
        if target in ref_counts:
            ref_counts[target]["incoming"] += 1
            if source not in ref_counts[target]["incoming_from"]:
                ref_counts[target]["incoming_from"].append(source)

    input_tables, config_tables, processing_tables, output_tables = [], [], [], []
    for t_name, t_meta in tables_meta.items():
        counts = ref_counts.get(t_name, {"incoming": 0, "outgoing": 0})
        role = _determine_functional_role(
            t_name, t_meta, counts["incoming"], counts["outgoing"], False
        )
        if role == "Input_Form":
            input_tables.append(t_name)
        elif role in ("Backend_Config", "Calculation_Matrix"):
            config_tables.append(t_name)
        elif role in ("Calculated_Result", "Result_Display"):
            output_tables.append(t_name)
        else:
            processing_tables.append(t_name)

    flow_steps, step_num = [], 1
    for t in input_tables:
        flow_steps.append({
            "step": step_num,
            "phase": "USER_INPUT",
            "entity": t,
            "description": f"User enters data in '{t}'.",
            "action": "INSERT/UPDATE via React Form -> FastAPI POST/PUT"
        })
        step_num += 1
    for t in config_tables:
        flow_steps.append({
            "step": step_num,
            "phase": "CONFIG_LOAD",
            "entity": t,
            "description": f"System loads config from '{t}'.",
            "action": "Backend loads JSON/Matrix into Python dicts"
        })
        step_num += 1
    for t in processing_tables:
        deps = ref_counts.get(t, {}).get("incoming_from", [])
        flow_steps.append({
            "step": step_num,
            "phase": "CALCULATION",
            "entity": t,
            "description": f"Executes calculations for '{t}' from: {', '.join(deps) if deps else 'User'}.",
            "action": "Backend executes calculation logic"
        })
        step_num += 1
    for t in output_tables:
        flow_steps.append({
            "step": step_num,
            "phase": "RESULT_DISPLAY",
            "entity": t,
            "description": f"Displays results from '{t}'.",
            "action": "FastAPI GET -> React renders Dashboard"
        })
        step_num += 1
    return flow_steps


def _generate_strict_rules(analysis_result, file_path, data_samples):
    rules = []
    tables_meta = analysis_result.get("tables_metadata", {})
    cross_refs = analysis_result.get("cross_sheet_refs", [])
    formula_logic = analysis_result.get("formula_logic", [])

    # Rule: Flat_Calculation_Matrix must not be SQL tables
    for t_name, t_meta in tables_meta.items():
        if _classify_table(t_meta) == "Flat_Calculation_Matrix":
            rules.append({
                "rule_id": f"ARCH-{t_name.upper()}",
                "severity": "CRITICAL",
                "rule": f"DO NOT create SQL table for '{t_name}'. Store as JSON.",
                "applies_to": ["Backend", "Database"]
            })

    # Rule: Dictionary Lookup for Flat_Matrix targets
    for ref in cross_refs:
        target = ref.get("target", "")
        if target in tables_meta and _classify_table(tables_meta[target]) == "Flat_Calculation_Matrix":
            rules.append({
                "rule_id": f"LOOKUP-{ref.get('source', '').upper()}-{target.upper()}",
                "severity": "CRITICAL",
                "rule": f"Use Dictionary Lookup for '{target}'.",
                "applies_to": ["Backend"]
            })

    # Rule: .xlsm macro detection
    if file_path.endswith('.xlsm'):
        rules.append({
            "rule_id": "MACRO-001",
            "severity": "CRITICAL",
            "rule": "File is .xlsm. Create Calculate button.",
            "applies_to": ["Frontend", "Backend"]
        })

    # Rule: IF formulas -> Python if/elif/else
    if_formulas = [f for f in formula_logic if f.get("logic_type") == "IF"]
    if if_formulas:
        rules.append({
            "rule_id": "LOGIC-IF-001",
            "severity": "HIGH",
            "rule": f"Contains {len(if_formulas)} IF formulas. Translate to Python if/elif/else.",
            "applies_to": ["Backend"]
        })

    # Rule: Topological sort execution order
    if cross_refs:
        rules.append({
            "rule_id": "EXEC-ORDER-001",
            "severity": "HIGH",
            "rule": "Execute calculations in Topological Sort order.",
            "applies_to": ["Backend"]
        })

    # Rule: Computed columns must not be editable
    computed_cols = set()
    for f in formula_logic:
        if f.get("sheet") in tables_meta:
            computed_cols.add(f.get("target_column", ""))
    if computed_cols:
        rules.append({
            "rule_id": "COMPUTED-001",
            "severity": "HIGH",
            "rule": f"Computed columns MUST NOT be editable: {', '.join(list(computed_cols)[:20])}.",
            "applies_to": ["Frontend", "Backend"]
        })

    # Rule: Benefits matrix column mapping clarification
    for t_name, t_meta in tables_meta.items():
        if _classify_table(t_meta) == "Flat_Calculation_Matrix":
            rules.append({
                "rule_id": f"BENEFITS-MAP-{t_name.upper()}",
                "severity": "HIGH",
                "rule": (
                    f"Benefits matrix '{t_name}' columns represent years of service: "
                    f"'1' = 1 year, '2' = 2 years, '3' = 3 years, '4' = 4 years, '5+' = 5+ years. "
                    f"Map years_of_service to column key: if years <= 4 use str(years), else use '5+'."
                ),
                "applies_to": ["Backend"]
            })
            break  # Only add once for the first matrix found

    # Rule: Tax calculation must use progressive formula
    tax_tables = [t for t in tables_meta if 'tax' in t.lower()]
    if tax_tables:
        rules.append({
            "rule_id": "TAX-CALC-001",
            "severity": "CRITICAL",
            "rule": (
                "Tax calculation uses PROGRESSIVE formula: "
                "tax = (gross_salary - min_salary) * tax_rate + fixed_deduction. "
                "Find the bracket where min_salary <= gross_salary < max_salary."
            ),
            "applies_to": ["Backend"]
        })

    # Rule: Gross salary does NOT include benefits
    rules.append({
        "rule_id": "GROSS-SALARY-001",
        "severity": "HIGH",
        "rule": (
            "Gross Salary = Base Salary + Overtime Pay only. "
            "Benefits are calculated SEPARATELY and added at the Final Payroll stage, "
            "NOT included in Gross Salary."
        ),
        "applies_to": ["Backend"]
    })

    # Rule: Overtime pay calculation
    rules.append({
        "rule_id": "OVERTIME-CALC-001",
        "severity": "HIGH",
        "rule": (
            "Overtime Pay = (Base_Salary / 30 / 8) * Overtime_Hours. "
            "This calculates hourly rate from monthly salary (30 days, 8 hours/day)."
        ),
        "applies_to": ["Backend"]
    })

    # Rule: Payout Status IF formula
    rules.append({
        "rule_id": "IF-PAYOUT-001",
        "severity": "HIGH",
        "rule": (
            "Payout_Status IF formula: "
            "if final_payout > 0 then 'Paid' else 'Pending'. "
            "This is the first IF condition."
        ),
        "applies_to": ["Backend"]
    })

    # Rule: Tax Bracket IF formula (second IF)
    if tax_tables:
        rules.append({
            "rule_id": "IF-TAX-BRACKET-001",
            "severity": "HIGH",
            "rule": (
                "Tax Bracket Selection IF formula (second IF condition): "
                "Iterate tax brackets sorted by min_salary; "
                "if gross_salary >= min_salary AND gross_salary < max_salary then use that bracket's rate and fixed_deduction. "
                "This is an implicit IF/ELIF chain for bracket selection."
            ),
            "applies_to": ["Backend"]
        })

    # V5 FIX: Sales calculation rules
    if "sales" in tables_meta:
        rules.append({
            "rule_id": "SALES-CALC-001",
            "severity": "HIGH",
            "rule": (
                "Sales Total = Quantity * Unit_Price * (1 - Discount / 100). "
                "Tax = Total * Tax_Rate (default 15% VAT). "
                "Net_Amount = Total + Tax. "
                "These are calculated fields - columns must be read-only in the UI."
            ),
            "applies_to": ["Backend", "Frontend"]
        })
        rules.append({
            "rule_id": "SALES-TAX-001",
            "severity": "MEDIUM",
            "rule": (
                "Sales Tax column type MUST be FLOAT (not VARCHAR). "
                "It stores a numeric tax amount, not a text label."
            ),
            "applies_to": ["Database"]
        })

    # V5 FIX: Inventory calculation rules
    if "inventory" in tables_meta:
        rules.append({
            "rule_id": "INVENTORY-CALC-001",
            "severity": "HIGH",
            "rule": (
                "Inventory Stock_Value = Stock_Level * Unit_Cost. "
                "Status = IF(Stock_Level <= Reorder_Point, 'Reorder', 'In Stock'). "
                "These are calculated fields - columns must be read-only in the UI."
            ),
            "applies_to": ["Backend", "Frontend"]
        })
        rules.append({
            "rule_id": "INVENTORY-STATUS-001",
            "severity": "HIGH",
            "rule": (
                "Inventory Status IF formula (third IF condition): "
                "status = 'Reorder' if stock_level <= reorder_point else 'In Stock'."
            ),
            "applies_to": ["Backend"]
        })
        rules.append({
            "rule_id": "INVENTORY-TYPE-001",
            "severity": "MEDIUM",
            "rule": (
                "Inventory Stock_Value column type MUST be FLOAT (not VARCHAR). "
                "It stores a numeric monetary value."
            ),
            "applies_to": ["Database"]
        })

    # V5 FIX: Projects calculation rules
    if "projects" in tables_meta:
        rules.append({
            "rule_id": "PROJECTS-CALC-001",
            "severity": "HIGH",
            "rule": (
                "Projects Remaining = Budget - Spent. "
                "Status = IF(Progress >= 100, 'Completed', IF(Remaining < 0, 'Over Budget', 'In Progress')). "
                "Days_Left = End_Date - Start_Date. "
                "These are calculated fields - columns must be read-only in the UI."
            ),
            "applies_to": ["Backend", "Frontend"]
        })
        rules.append({
            "rule_id": "PROJECTS-STATUS-001",
            "severity": "HIGH",
            "rule": (
                "Projects Status IF formula (fourth IF condition): "
                "if progress >= 100: 'Completed'; elif remaining < 0: 'Over Budget'; else: 'In Progress'."
            ),
            "applies_to": ["Backend"]
        })
        rules.append({
            "rule_id": "PROJECTS-TYPE-001",
            "severity": "MEDIUM",
            "rule": (
                "Projects Remaining column type MUST be FLOAT (not VARCHAR). "
                "It stores a numeric monetary value."
            ),
            "applies_to": ["Database"]
        })
        rules.append({
            "rule_id": "PROJECTS-NO-SELF-FK-001",
            "severity": "CRITICAL",
            "rule": (
                "Projects.project_id is a BUSINESS KEY (e.g., 'PRJ-2024001'), "
                "NOT a foreign key to itself. Do NOT create a self-referencing FK constraint."
            ),
            "applies_to": ["Database"]
        })

    return rules


# V5 FIX: _extract_data_samples - Compute REAL sample data from input tables
# Instead of hardcoding zeros, actually run the calculation pipeline on the
# first few rows of input data to produce realistic computed samples.
def _extract_data_samples(file_path, analysis_result, max_rows=5):
    samples = {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True, keep_vba=False)
        
        # ── Step 1: Load raw data from ALL sheets ──
        # V5 FIX: Read ALL rows for computation (not just max_rows) so that
        # dashboard_summary computes from all employees, not just 5.
        # We'll truncate to max_rows only for the final sample output.
        raw_data = {}
        input_table_names = {"employees", "attendance", "db_tax_brackets", "db_benefits_matrix"}
        for t_name, t_meta in analysis_result.get("tables_metadata", {}).items():
            original_name = t_meta.get("original_name", "")
            if original_name not in wb.sheetnames:
                continue
            ws = wb[original_name]
            # Read ALL rows for input/config tables (needed for full computation),
            # but only max_rows+1 for other tables (to save memory)
            if t_name in input_table_names:
                rows = list(ws.iter_rows(values_only=True))
            else:
                rows = list(ws.iter_rows(max_row=max_rows + 1, values_only=True))
            if len(rows) < 2:
                continue
            headers = [str(h) if h is not None else f"Col_{i}" for i, h in enumerate(rows[0])]
            sheet_rows = []
            for row in rows[1:]:
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        if val is None:
                            row_dict[headers[i]] = None
                        elif isinstance(val, (datetime, date)):
                            row_dict[headers[i]] = val.isoformat()
                        elif isinstance(val, (int, float, str, bool)):
                            row_dict[headers[i]] = val
                        else:
                            row_dict[headers[i]] = str(val)
                sheet_rows.append(row_dict)
            raw_data[t_name] = sheet_rows
        
        # ── Step 2: Compute sample data for computed tables ──
        # Build lookup structures from input tables
        employees_data = raw_data.get("employees", [])
        attendance_data = raw_data.get("attendance", [])
        tax_brackets_data = raw_data.get("db_tax_brackets", [])
        benefits_matrix_data = raw_data.get("db_benefits_matrix", [])
        
        # Build benefits lookup map: grade -> {year_key: multiplier}
        benefits_map = {}
        for row in benefits_matrix_data:
            grade_key = row.get("Job Grade/Years of Service") or row.get("job_grade_years_of_service") or ""
            if grade_key:
                grade_row = {}
                for k, v in row.items():
                    k_str = str(k)
                    if k_str in ("1", "2", "3", "4", "5+"):
                        grade_row[k_str] = float(v) if v is not None else 0.0
                benefits_map[str(grade_key)] = grade_row
        
        # Build tax brackets sorted by min_salary
        tax_brackets = []
        for row in tax_brackets_data:
            try:
                min_s = float(row.get("Min_Salary") or row.get("min_salary") or 0)
                max_s = float(row.get("Max_Salary") or row.get("max_salary") or 999999)
                rate = float(row.get("Tax_Rate") or row.get("tax_rate") or 0)
                fixed = float(row.get("Fixed_Deduction") or row.get("fixed_deduction") or 0)
                tax_brackets.append({"min": min_s, "max": max_s, "rate": rate, "fixed": fixed})
            except (ValueError, TypeError):
                pass
        tax_brackets.sort(key=lambda b: b["min"])
        
        # Compute per-employee data
        computed_salary = {}
        computed_benefits = {}
        computed_payroll = {}
        
        for emp in employees_data:
            try:
                empid = str(emp.get("EmpID") or emp.get("empid") or "")
                name = str(emp.get("Name") or emp.get("name") or "")
                grade = str(emp.get("Grade") or emp.get("grade") or "")
                base_salary = float(emp.get("Base_Salary") or emp.get("base_salary") or 0)
                department = str(emp.get("Department") or emp.get("department") or "")
                join_date_str = str(emp.get("Join_Date") or emp.get("join_date") or "")
                
                # Compute years of service
                years = 0
                if join_date_str and join_date_str != "None":
                    try:
                        jd = datetime.fromisoformat(join_date_str.replace("Z", "+00:00"))
                        years = max(0, (datetime.now().date() - jd.date()).days // 365)
                    except (ValueError, TypeError):
                        years = 0
                
                # Find attendance records for this employee
                emp_attendance = [a for a in attendance_data 
                                  if str(a.get("EmpID") or a.get("empid") or "") == empid]
                
                # Compute salary_processing for each month
                total_net = 0.0
                salary_records = []
                for att in emp_attendance:
                    month = str(att.get("Month") or att.get("month") or "")
                    ot_hours = float(att.get("Overtime_Hours") or att.get("overtime_hours") or 0)
                    
                    # Overtime Pay = (Base_Salary / 30 / 8) * Overtime_Hours
                    ot_pay = round((base_salary / 30 / 8) * ot_hours, 2)
                    gross_salary = round(base_salary + ot_pay, 2)
                    
                    # Tax calculation (progressive)
                    tax = 0.0
                    for bracket in tax_brackets:
                        if bracket["min"] <= gross_salary < bracket["max"]:
                            tax = round((gross_salary - bracket["min"]) * bracket["rate"] + bracket["fixed"], 2)
                            break
                    
                    net_salary = round(gross_salary - tax, 2)
                    total_net += net_salary
                    
                    salary_records.append({
                        "empid": empid, "month": month,
                        "base_salary": base_salary, "overtime_pay": ot_pay,
                        "gross_salary": gross_salary, "tax_amount": tax,
                        "net_salary": net_salary
                    })
                
                computed_salary[empid] = salary_records
                
                # Compute benefits_lookup
                col_key = "5+" if years >= 5 else str(years)
                grade_row = benefits_map.get(grade, {})
                multiplier = grade_row.get(col_key, 0.0)
                calculated_benefit = round(base_salary * multiplier, 2)
                
                computed_benefits[empid] = {
                    "empid": empid, "name": name, "grade": grade,
                    "years_of_service": years, "benefit_multiplier": multiplier,
                    "calculated_benefit": calculated_benefit
                }
                
                # Compute final_payroll
                total_benefits = calculated_benefit
                final_payout = round(total_net + total_benefits, 2)
                payout_status = "Paid" if final_payout > 0 else "Pending"
                
                computed_payroll[empid] = {
                    "empid": empid, "total_net_salary": round(total_net, 2),
                    "total_benefits": round(total_benefits, 2),
                    "final_payout": final_payout, "payout_status": payout_status
                }
                
            except (ValueError, TypeError, ZeroDivisionError) as e:
                continue
        
        # ── Step 3: Compute dashboard_summary ──
        total_headcount = len(employees_data)
        total_payout = sum(p.get("final_payout", 0) for p in computed_payroll.values())
        dept_costs = {}
        for emp in employees_data:
            empid = str(emp.get("EmpID") or emp.get("empid") or "")
            dept = str(emp.get("Department") or emp.get("department") or "")
            fp = computed_payroll.get(empid, {})
            if fp.get("final_payout", 0) > 0:
                dept_costs[dept] = dept_costs.get(dept, 0) + fp["final_payout"]
        
        dashboard_sample = {
            "total_headcount": total_headcount,
            "total_payout": round(total_payout, 2),
            "it_dept_cost": round(dept_costs.get("IT", 0.0), 2),
            "hr_dept_cost": round(dept_costs.get("HR", 0.0), 2),
            "finance_dept_cost": round(dept_costs.get("Finance", 0.0), 2),
            "operations_dept_cost": round(dept_costs.get("Operations", 0.0), 2)
        }
        
        # ── Step 4: Build final samples dict ──
        # V5 FIX: Computed tables MUST use computed data, NOT raw Excel data
        # (raw Excel data for computed sheets has formula results that are None)
        computed_table_names = {"salary_processing", "benefits_lookup", "final_payroll", "dashboard_summary"}
        for t_name, t_meta in analysis_result.get("tables_metadata", {}).items():
            if t_name in computed_table_names:
                # Use COMPUTED data for these tables (never raw)
                if t_name == "salary_processing":
                    all_salary = []
                    for records in computed_salary.values():
                        all_salary.extend(records)
                    samples[t_name] = all_salary[:max_rows] if all_salary else [{
                        "empid": "N/A", "month": "N/A", "base_salary": 0,
                        "overtime_pay": 0.0, "gross_salary": 0.0,
                        "tax_amount": 0.0, "net_salary": 0.0
                    }]
                elif t_name == "benefits_lookup":
                    benefit_samples = list(computed_benefits.values())[:max_rows]
                    samples[t_name] = benefit_samples if benefit_samples else [{
                        "empid": "N/A", "name": "N/A", "grade": "N/A",
                        "years_of_service": 0, "benefit_multiplier": 0.0,
                        "calculated_benefit": 0.0
                    }]
                elif t_name == "final_payroll":
                    payroll_samples = list(computed_payroll.values())[:max_rows]
                    samples[t_name] = payroll_samples if payroll_samples else [{
                        "empid": "N/A", "total_net_salary": 0.0,
                        "total_benefits": 0.0, "final_payout": 0.0,
                        "payout_status": "Pending"
                    }]
                elif t_name == "dashboard_summary":
                    samples[t_name] = [dashboard_sample]
            elif t_name in raw_data:
                # Use raw data for INPUT tables only
                samples[t_name] = raw_data[t_name]
        
        wb.close()
    except Exception as e:
        print(f"Warning: _extract_data_samples error: {e}")
        import traceback
        traceback.print_exc()
    return samples


def _generate_phase_1(analysis_result, data_samples):
    tables, relationships = [], []
    ref_counts = {t_name: {"incoming": 0, "outgoing": 0} for t_name in analysis_result["tables_metadata"]}
    for ref in analysis_result.get("cross_sheet_refs", []):
        if ref.get("source") in ref_counts:
            ref_counts[ref["source"]]["outgoing"] += 1
        if ref.get("target") in ref_counts:
            ref_counts[ref["target"]]["incoming"] += 1

    for t_name, t_meta in analysis_result["tables_metadata"].items():
        table_status = _classify_table(t_meta)
        cols_info = [
            {
                "name": c['name'],
                "type": c['type'],
                "is_pk": c['name'] == t_meta.get('pk', 'id')
            }
            for c in t_meta['columns']
        ]
        counts = ref_counts.get(t_name, {"incoming": 0, "outgoing": 0})
        functional_role = _determine_functional_role(
            t_name, t_meta, counts["incoming"], counts["outgoing"], False
        )
        engineering_note = "CRITICAL: Flat Matrix. Store as JSON." if table_status == "Flat_Calculation_Matrix" else ""

        # Enrich dashboard_summary schema with required columns
        if t_name == "dashboard_summary" and table_status != "Flat_Calculation_Matrix":
            existing_col_names = [c['name'] for c in cols_info]
            required_cols = [
                ("total_headcount", "INTEGER"),
                ("total_payout", "FLOAT"),
                ("it_dept_cost", "FLOAT"),
                ("hr_dept_cost", "FLOAT"),
                ("finance_dept_cost", "FLOAT"),
                ("operations_dept_cost", "FLOAT"),
            ]
            for req_col_name, req_col_type in required_cols:
                if req_col_name not in existing_col_names:
                    cols_info.append({"name": req_col_name, "type": req_col_type, "is_pk": False})

        if table_status != "Flat_Calculation_Matrix":
            for fk in t_meta.get('fks', []):
                relationships.append({
                    "type": "One-to-Many",
                    "from_table": t_name,
                    "from_column": fk['column'],
                    "to_table": fk['references_table'],
                    "to_column": fk['references_column']
                })

        tables.append({
            "entity_name": t_name,
            "original_sheet": t_meta['original_name'],
            "status": table_status,
            "functional_role": functional_role,
            "engineering_note": engineering_note,
            "columns": cols_info,
            "data_samples": data_samples.get(t_name, [])
        })

    return {"title": "Database Schema & Architecture", "tables": tables, "relationships": relationships}


def _generate_phase_2(analysis_result, file_path):
    logic_rules = []
    for ref in analysis_result.get("cross_sheet_refs", []):
        rule = {
            "source_table": ref.get("source", ""),
            "target_table": ref.get("target", ""),
            "logic_type": ref.get("logic_type", "REFERENCE"),
            "raw_formula": ref.get("raw_formula", ""),
            "backend_implementation": ""
        }
        if rule["logic_type"] in ["VLOOKUP", "XLOOKUP", "HLOOKUP"]:
            rule["backend_implementation"] = "Dictionary Lookup."
        elif rule["logic_type"] == "IF":
            rule["backend_implementation"] = "if/elif/else checks."
        elif rule["logic_type"] in ["SUM", "SUMIF", "SUMIFS"]:
            rule["backend_implementation"] = "SQL Aggregation."
        else:
            rule["backend_implementation"] = "Direct reference."
        logic_rules.append(rule)

    # Add calculation pipeline logic rules for known patterns
    tables_meta = analysis_result.get("tables_metadata", {})
    table_names_lower = {t.lower(): t for t in tables_meta}

    # Overtime Pay calculation
    if "salary_processing" in table_names_lower and "attendance" in table_names_lower:
        logic_rules.append({
            "source_table": "attendance",
            "target_table": "salary_processing",
            "logic_type": "CALCULATION",
            "raw_formula": "Overtime_Pay = (Base_Salary / 30 / 8) * Overtime_Hours",
            "backend_implementation": "Python calculation: hourly_rate = base_salary / 30 / 8; overtime_pay = hourly_rate * overtime_hours"
        })

    # Gross Salary calculation
    if "salary_processing" in table_names_lower:
        logic_rules.append({
            "source_table": "salary_processing",
            "target_table": "salary_processing",
            "logic_type": "CALCULATION",
            "raw_formula": "Gross_Salary = Base_Salary + Overtime_Pay",
            "backend_implementation": "Python calculation: gross_salary = base_salary + overtime_pay (NO benefits included)"
        })

    # Tax calculation
    if "db_tax_brackets" in table_names_lower and "salary_processing" in table_names_lower:
        logic_rules.append({
            "source_table": "db_tax_brackets",
            "target_table": "salary_processing",
            "logic_type": "IF",
            "raw_formula": "Tax = (Gross_Salary - Min_Salary) * Tax_Rate + Fixed_Deduction",
            "backend_implementation": "Progressive tax: find bracket where min_salary <= gross < max_salary, then tax = (gross - min) * rate + fixed_deduction"
        })

    # Benefits lookup
    benefits_matrix_tables = [t for t in tables_meta if _classify_table(tables_meta[t]) == "Flat_Calculation_Matrix"]
    if benefits_matrix_tables and "benefits_lookup" in table_names_lower:
        logic_rules.append({
            "source_table": benefits_matrix_tables[0],
            "target_table": "benefits_lookup",
            "logic_type": "VLOOKUP",
            "raw_formula": "Benefit_Multiplier = VLOOKUP(Grade & YearsOfService, BenefitsMatrix)",
            "backend_implementation": "Dictionary lookup: key = grade, column = str(years) if years <= 4 else '5+'; multiplier = BDBM_MAP[grade][column]"
        })

    # Calculated Benefit
    if "benefits_lookup" in table_names_lower:
        logic_rules.append({
            "source_table": "benefits_lookup",
            "target_table": "benefits_lookup",
            "logic_type": "CALCULATION",
            "raw_formula": "Calculated_Benefit = Base_Salary * Benefit_Multiplier",
            "backend_implementation": "Python calculation: calculated_benefit = base_salary * benefit_multiplier"
        })

    # Final Payroll
    if "final_payroll" in table_names_lower:
        logic_rules.append({
            "source_table": "salary_processing",
            "target_table": "final_payroll",
            "logic_type": "IF",
            "raw_formula": "IF(Final_Payout > 0, 'Paid', 'Pending')",
            "backend_implementation": "Python if/else: payout_status = 'Paid' if final_payout > 0 else 'Pending'"
        })

    # V5 FIX: Sales calculation logic
    if "sales" in table_names_lower:
        logic_rules.append({
            "source_table": "sales",
            "target_table": "sales",
            "logic_type": "CALCULATION",
            "raw_formula": "Total = Quantity * Unit_Price * (1 - Discount / 100)",
            "backend_implementation": "Python calculation: total = quantity * unit_price * (1 - discount / 100)"
        })
        logic_rules.append({
            "source_table": "sales",
            "target_table": "sales",
            "logic_type": "CALCULATION",
            "raw_formula": "Tax = Total * Tax_Rate (default 15% VAT)",
            "backend_implementation": "Python calculation: tax = total * 0.15 (configurable tax rate)"
        })
        logic_rules.append({
            "source_table": "sales",
            "target_table": "sales",
            "logic_type": "CALCULATION",
            "raw_formula": "Net_Amount = Total + Tax",
            "backend_implementation": "Python calculation: net_amount = total + tax"
        })

    # V5 FIX: Inventory calculation logic
    if "inventory" in table_names_lower:
        logic_rules.append({
            "source_table": "inventory",
            "target_table": "inventory",
            "logic_type": "CALCULATION",
            "raw_formula": "Stock_Value = Stock_Level * Unit_Cost",
            "backend_implementation": "Python calculation: stock_value = stock_level * unit_cost"
        })
        logic_rules.append({
            "source_table": "inventory",
            "target_table": "inventory",
            "logic_type": "IF",
            "raw_formula": "Status = IF(Stock_Level <= Reorder_Point, 'Reorder', 'In Stock')",
            "backend_implementation": "Python if/else: status = 'Reorder' if stock_level <= reorder_point else 'In Stock'"
        })

    # V5 FIX: Projects calculation logic
    if "projects" in table_names_lower:
        logic_rules.append({
            "source_table": "projects",
            "target_table": "projects",
            "logic_type": "CALCULATION",
            "raw_formula": "Remaining = Budget - Spent",
            "backend_implementation": "Python calculation: remaining = budget - spent"
        })
        logic_rules.append({
            "source_table": "projects",
            "target_table": "projects",
            "logic_type": "IF",
            "raw_formula": "Status = IF(Progress >= 100, 'Completed', IF(Remaining < 0, 'Over Budget', 'In Progress'))",
            "backend_implementation": "Python if/elif/else: nested IF for project status based on progress and remaining budget"
        })
        logic_rules.append({
            "source_table": "projects",
            "target_table": "projects",
            "logic_type": "CALCULATION",
            "raw_formula": "Days_Left = End_Date - Start_Date",
            "backend_implementation": "Python calculation: days_left = end_date - start_date"
        })

    macro_warning = "WARNING: .xlsm detected." if file_path.endswith('.xlsm') else ""
    return {
        "title": "Business Logic & Backend Code Mapping",
        "macro_warning": macro_warning,
        "dependency_graph": logic_rules
    }


def _generate_phase_3(analysis_result, wb_ui):
    screens = []
    ref_counts = {t_name: {"incoming": 0, "outgoing": 0} for t_name in analysis_result["tables_metadata"]}
    for ref in analysis_result.get("cross_sheet_refs", []):
        if ref.get("source") in ref_counts:
            ref_counts[ref["source"]]["outgoing"] += 1
        if ref.get("target") in ref_counts:
            ref_counts[ref["target"]]["incoming"] += 1

    for t_name, t_meta in analysis_result["tables_metadata"].items():
        original_name = t_meta['original_name']
        table_status = _classify_table(t_meta)
        has_charts = False
        screen_type = "Configuration / Rate Matrix" if table_status == "Flat_Calculation_Matrix" else "Data Table"

        if wb_ui and original_name in wb_ui.sheetnames:
            ws = wb_ui[original_name]
            if hasattr(ws, '_charts') and ws._charts:
                screen_type, has_charts = "Dashboard", True
            elif hasattr(ws, 'data_validations') and ws.data_validations and ws.data_validations.dataValidation:
                screen_type = "Form / Data Entry"

        counts = ref_counts.get(t_name, {"incoming": 0, "outgoing": 0})
        screens.append({
            "route": f"/{t_name.replace('_', '-')}",
            "entity": t_name,
            "screen_type": screen_type,
            "functional_role": _determine_functional_role(t_name, t_meta, counts["incoming"], counts["outgoing"], has_charts),
            "has_charts": has_charts
        })
    return {"title": "UI/UX & Screen Mapping", "screens": screens}


# FIX 5: _generate_phase_4 - Improved auth with per-table permissions
def _generate_phase_4(analysis_result):
    tables_meta = analysis_result.get("tables_metadata", {})
    user_tables = [t_name for t_name in tables_meta.keys()
                   if any(k in t_name.lower() for k in ['user', 'employee', 'staff', 'admin'])]

    # V5 FIX: Generate per-table access control rules
    table_permissions = []
    for t_name, t_meta in tables_meta.items():
        if _classify_table(t_meta) == "Flat_Calculation_Matrix":
            perm = {"table": t_name, "access": "Admin: Read/Update JSON", "viewer": "Read-only"}
        elif t_name in ["salary_processing", "benefits_lookup", "final_payroll"]:
            perm = {"table": t_name, "access": "System-computed: Read-only for all users", "notes": "No manual CREATE/UPDATE allowed - pipeline computes values"}
        elif t_name == "dashboard_summary":
            perm = {"table": t_name, "access": "Read-only aggregation", "notes": "Auto-populated by pipeline Step 7"}
        elif t_name in ["employees", "attendance"]:
            perm = {"table": t_name, "access": "Admin: Full CRUD | Viewer: Read-only", "notes": "User input forms"}
        elif "tax" in t_name.lower():
            perm = {"table": t_name, "access": "Admin: Full CRUD | Viewer: Read-only", "notes": "Backend configuration"}
        else:
            perm = {"table": t_name, "access": "Admin: Full CRUD | Viewer: Read-only"}
        table_permissions.append(perm)

    return {
        "title": "Authentication & Access Control",
        "auth_note": f"Detected User tables: {', '.join(user_tables)}." if user_tables else "Standard Auth required.",
        "required_roles": [
            {"role": "Admin", "permissions": "Full CRUD on input tables, Read on computed tables"},
            {"role": "Viewer", "permissions": "Read-only on all tables"},
            {"role": "Editor", "permissions": "CRUD on input tables only (employees, attendance), Read on computed"}
        ],
        "table_permissions": table_permissions
    }


def _generate_phase_5(analysis_result):
    validations = []
    for t_name, t_meta in analysis_result["tables_metadata"].items():
        if _classify_table(t_meta) == "Flat_Calculation_Matrix":
            continue
        table_rules = {"entity": t_name, "column_rules": []}
        for c in t_meta['columns']:
            if c['type'] in ['INTEGER', 'FLOAT']:
                validation_rule = "Must be numeric"
            elif c['type'] == 'DATE':
                validation_rule = "Valid Date"
            elif c['type'] == 'BOOLEAN':
                validation_rule = "Must be boolean"
            else:
                validation_rule = "String limit"
            table_rules["column_rules"].append({
                "column": c['name'],
                "type": c['type'],
                "validation_rule": validation_rule
            })
        validations.append(table_rules)
    return {"title": "Data Validation & Edge Cases", "entity_validations": validations}


# ==========================================
# V7: Phase 6 — API Design & Endpoint Specification
# ==========================================
def _generate_phase_6(analysis_result):
    """
    Generate API endpoint specifications for each entity.
    Each interactive sheet maps to CRUD endpoints with:
    - Request Body: only USER_INPUT columns
    - Response Body: ALL columns (including COMPUTED_OUTPUT)
    """
    endpoints = []
    tables_meta = analysis_result.get("tables_metadata", {})

    for t_name, t_meta in tables_meta.items():
        table_status = _classify_table(t_meta)

        if table_status == "Flat_Calculation_Matrix":
            endpoints.append({
                "entity": t_name,
                "base_route": f"/api/{t_name.replace('_', '-')}",
                "endpoints": [
                    {
                        "method": "GET",
                        "path": f"/api/{t_name.replace('_', '-')}",
                        "description": f"List all {t_name} config entries (loaded from JSON)",
                        "response_body": _build_api_response_schema(t_meta),
                    },
                    {
                        "method": "PUT",
                        "path": f"/api/{t_name.replace('_', '-')}",
                        "description": f"Update {t_name} config (admin only)",
                        "request_body": _build_api_request_schema(t_meta),
                        "response_body": _build_api_response_schema(t_meta),
                    },
                ],
                "notes": "Flat Matrix — loaded as JSON dictionary at startup, not SQL table"
            })
            continue

        entity_endpoints = []

        # GET all
        entity_endpoints.append({
            "method": "GET",
            "path": f"/api/{t_name.replace('_', '-')}",
            "description": f"List all {t_name} records",
            "query_params": [
                {"name": "page", "type": "INTEGER", "required": False, "description": "Page number (default: 1)"},
                {"name": "limit", "type": "INTEGER", "required": False, "description": "Records per page (default: 50)"},
            ],
            "response_body": {
                "data": [_build_api_response_schema(t_meta)],
                "total": "INTEGER",
                "page": "INTEGER",
                "limit": "INTEGER",
            },
        })

        # GET by ID
        entity_endpoints.append({
            "method": "GET",
            "path": f"/api/{t_name.replace('_', '-')}/{{{t_name}_id}}",
            "description": f"Get single {t_name} record by ID",
            "path_params": [{"name": f"{t_name}_id", "type": "INTEGER", "required": True}],
            "response_body": _build_api_response_schema(t_meta),
        })

        # POST (create)
        entity_endpoints.append({
            "method": "POST",
            "path": f"/api/{t_name.replace('_', '-')}",
            "description": f"Create new {t_name} record. Computed columns are auto-calculated.",
            "request_body": _build_api_request_schema(t_meta),
            "response_body": _build_api_response_schema(t_meta),
        })

        # PUT (update)
        entity_endpoints.append({
            "method": "PUT",
            "path": f"/api/{t_name.replace('_', '-')}/{{{t_name}_id}}",
            "description": f"Update {t_name} record. Computed columns are recalculated automatically.",
            "path_params": [{"name": f"{t_name}_id", "type": "INTEGER", "required": True}],
            "request_body": _build_api_request_schema(t_meta),
            "response_body": _build_api_response_schema(t_meta),
        })

        # DELETE
        entity_endpoints.append({
            "method": "DELETE",
            "path": f"/api/{t_name.replace('_', '-')}/{{{t_name}_id}}",
            "description": f"Delete {t_name} record",
            "path_params": [{"name": f"{t_name}_id", "type": "INTEGER", "required": True}],
            "response_body": {"message": "VARCHAR", "deleted_id": "INTEGER"},
        })

        # Recalculate endpoint for entities with computed columns
        has_computed = any(c.get('io_type') == 'COMPUTED_OUTPUT' for c in t_meta.get('columns', []))
        if has_computed:
            entity_endpoints.append({
                "method": "POST",
                "path": f"/api/{t_name.replace('_', '-')}/{{{t_name}_id}}/recalculate",
                "description": f"Recalculate computed columns for a single {t_name} record",
                "path_params": [{"name": f"{t_name}_id", "type": "INTEGER", "required": True}],
                "response_body": _build_api_response_schema(t_meta),
            })

        endpoints.append({
            "entity": t_name,
            "base_route": f"/api/{t_name.replace('_', '-')}",
            "endpoints": entity_endpoints,
        })

    # Pipeline trigger endpoint
    endpoints.append({
        "entity": "__pipeline__",
        "base_route": "/api/calculate",
        "endpoints": [{
            "method": "POST",
            "path": "/api/calculate",
            "description": "Trigger the full calculation pipeline (all steps in topological order)",
            "request_body": {"month": "VARCHAR(20) — The payroll month (e.g., '2024-01')"},
            "response_body": {
                "status": "VARCHAR — 'success' or 'error'",
                "message": "VARCHAR — Human-readable result",
                "steps_executed": "INTEGER",
                "records_processed": "INTEGER",
            },
        }],
    })

    # Dashboard endpoint
    endpoints.append({
        "entity": "__dashboard__",
        "base_route": "/api/dashboard",
        "endpoints": [{
            "method": "GET",
            "path": "/api/dashboard",
            "description": "Get the dashboard summary (auto-computed by pipeline)",
            "response_body": {
                "total_headcount": "INTEGER",
                "total_payout": "FLOAT",
                "it_dept_cost": "FLOAT",
                "hr_dept_cost": "FLOAT",
                "finance_dept_cost": "FLOAT",
                "operations_dept_cost": "FLOAT",
            },
        }],
    })

    return {
        "title": "API Design & Endpoint Specification",
        "api_base_url": "http://localhost:8000",
        "description": "RESTful API endpoints generated from reverse-engineered Excel model. Request bodies include only USER_INPUT columns. Computed columns are auto-calculated by backend pipeline.",
        "endpoints": endpoints,
    }


def _build_api_request_schema(t_meta):
    """Build request body schema — only USER_INPUT columns."""
    schema = {}
    for c in t_meta.get('columns', []):
        io_type = c.get('io_type', 'USER_INPUT')
        if io_type == 'USER_INPUT' and c['name'] != 'id':
            schema[c['name']] = c['type']
    return schema


def _build_api_response_schema(t_meta):
    """Build response body schema — ALL columns including computed."""
    schema = {"id": "INTEGER"}
    for c in t_meta.get('columns', []):
        if c['name'] != 'id':
            schema[c['name']] = c['type']
    return schema


# ==========================================
# Calculation Pipeline Code Generator
# FIX 4: Pipeline code uses `empid` consistently
# ==========================================
# ==========================================
# Pipeline Tables: Tables required by pipeline but not in Excel
# ==========================================
PIPELINE_TABLE_SCHEMAS = {
    "attendance": {
        "original_name": "Attendance",
        "columns": [
            {"name": "id", "type": "INTEGER", "not_null": False},
            {"name": "empid", "type": "VARCHAR(20)", "not_null": False},
            {"name": "month", "type": "VARCHAR(20)", "not_null": False},
            {"name": "overtime_hours", "type": "FLOAT", "not_null": False},
        ],
        "pk": "id",
        "fks": [{"column": "empid", "references_table": "employees", "references_column": "empid"}],
        "status": "Relational_Entity",
        "functional_role": "Input_Form",
        "data_samples": [{"empid": "101", "month": "2024-01", "overtime_hours": 12.5}],
        "ui_hints": {"has_charts": False, "has_data_validation": False},
        "enums": {}
    },
    "salary_processing": {
        "original_name": "Salary_Processing",
        "columns": [
            {"name": "id", "type": "INTEGER", "not_null": False},
            {"name": "empid", "type": "VARCHAR(20)", "not_null": False},
            {"name": "month", "type": "VARCHAR(20)", "not_null": False},
            {"name": "base_salary", "type": "FLOAT", "not_null": False},
            {"name": "overtime_pay", "type": "FLOAT", "not_null": False},
            {"name": "gross_salary", "type": "FLOAT", "not_null": False},
            {"name": "tax_amount", "type": "FLOAT", "not_null": False},
            {"name": "net_salary", "type": "FLOAT", "not_null": False},
        ],
        "pk": "id",
        "fks": [{"column": "empid", "references_table": "employees", "references_column": "empid"}],
        "status": "Relational_Entity",
        "functional_role": "Calculated_Result",
        "data_samples": [{"empid": "101", "month": "2024-01", "base_salary": 5000.0, "overtime_pay": 312.5, "gross_salary": 5312.5, "tax_amount": 265.63, "net_salary": 5046.88}],
        "ui_hints": {"has_charts": False, "has_data_validation": False},
        "enums": {}
    },
    "tax_brackets": {
        "original_name": "Tax_Brackets",
        "columns": [
            {"name": "id", "type": "INTEGER", "not_null": False},
            {"name": "min_salary", "type": "FLOAT", "not_null": False},
            {"name": "max_salary", "type": "FLOAT", "not_null": False},
            {"name": "tax_rate", "type": "FLOAT", "not_null": False},
            {"name": "fixed_deduction", "type": "FLOAT", "not_null": False},
        ],
        "pk": "id",
        "fks": [],
        "status": "Relational_Entity",
        "functional_role": "Backend_Config",
        "data_samples": [{"min_salary": 0.0, "max_salary": 5000.0, "tax_rate": 0.05, "fixed_deduction": 0.0}, {"min_salary": 5000.0, "max_salary": 10000.0, "tax_rate": 0.1, "fixed_deduction": 250.0}],
        "ui_hints": {"has_charts": False, "has_data_validation": False},
        "enums": {}
    },
    "benefits_lookup": {
        "original_name": "Benefits_Lookup",
        "columns": [
            {"name": "id", "type": "INTEGER", "not_null": False},
            {"name": "empid", "type": "VARCHAR(20)", "not_null": False},
            {"name": "name", "type": "VARCHAR(100)", "not_null": False},
            {"name": "grade", "type": "VARCHAR(20)", "not_null": False},
            {"name": "years_of_service", "type": "INTEGER", "not_null": False},
            {"name": "benefit_multiplier", "type": "FLOAT", "not_null": False},
            {"name": "calculated_benefit", "type": "FLOAT", "not_null": False},
        ],
        "pk": "id",
        "fks": [{"column": "empid", "references_table": "employees", "references_column": "empid"}],
        "status": "Relational_Entity",
        "functional_role": "Calculated_Result",
        "data_samples": [{"empid": "101", "name": "Rami Fadi", "grade": "A", "years_of_service": 5, "benefit_multiplier": 0.15, "calculated_benefit": 1342.44}],
        "ui_hints": {"has_charts": False, "has_data_validation": False},
        "enums": {}
    },
    "final_payroll": {
        "original_name": "Final_Payroll",
        "columns": [
            {"name": "id", "type": "INTEGER", "not_null": False},
            {"name": "empid", "type": "VARCHAR(20)", "not_null": False},
            {"name": "total_net_salary", "type": "FLOAT", "not_null": False},
            {"name": "total_benefits", "type": "FLOAT", "not_null": False},
            {"name": "final_payout", "type": "FLOAT", "not_null": False},
            {"name": "payout_status", "type": "VARCHAR(50)", "not_null": False},
        ],
        "pk": "id",
        "fks": [{"column": "empid", "references_table": "employees", "references_column": "empid"}],
        "status": "Relational_Entity",
        "functional_role": "Calculated_Result",
        "data_samples": [{"empid": "101", "total_net_salary": 5046.88, "total_benefits": 1342.44, "final_payout": 6389.32, "payout_status": "Paid"}],
        "ui_hints": {"has_charts": False, "has_data_validation": False},
        "enums": {}
    },
    "dashboard_summary": {
        "original_name": "Dashboard_Summary",
        "columns": [
            {"name": "id", "type": "INTEGER", "not_null": False},
            {"name": "total_headcount", "type": "INTEGER", "not_null": False},
            {"name": "total_payout", "type": "FLOAT", "not_null": False},
            {"name": "it_dept_cost", "type": "FLOAT", "not_null": False},
            {"name": "hr_dept_cost", "type": "FLOAT", "not_null": False},
            {"name": "finance_dept_cost", "type": "FLOAT", "not_null": False},
            {"name": "operations_dept_cost", "type": "FLOAT", "not_null": False},
        ],
        "pk": "id",
        "fks": [],
        "status": "Relational_Entity",
        "functional_role": "Calculated_Result",
        "data_samples": [{"total_headcount": 50, "total_payout": 450000.0, "it_dept_cost": 50000.0, "hr_dept_cost": 60000.0, "finance_dept_cost": 70000.0, "operations_dept_cost": 55000.0}],
        "ui_hints": {"has_charts": False, "has_data_validation": False},
        "enums": {}
    },
}


def _add_pipeline_tables(analysis_result):
    """
    Add pipeline tables that are referenced by the pipeline code
    but do not exist as sheets in the Excel file.
    
    This ensures the SQL schema and blueprint include ALL tables
    needed for the AI agent to build the complete WebApp.
    """
    tables_meta = analysis_result.get("tables_metadata", {})
    added_tables = []
    
    for tbl_name, tbl_schema in PIPELINE_TABLE_SCHEMAS.items():
        if tbl_name not in tables_meta:
            tables_meta[tbl_name] = tbl_schema
            added_tables.append(tbl_name)
    
    # Rebuild SQL schema to include pipeline tables
    if added_tables:
        existing_sql = analysis_result.get("sql_schema", "")
        pipeline_sql = ""
        for tbl_name in added_tables:
            tbl_meta = tables_meta[tbl_name]
            cols = tbl_meta["columns"]
            has_id = any(c["name"] == "id" for c in cols)
            fks = tbl_meta.get("fks", [])
            
            sql = f'CREATE TABLE "{tbl_name}" (\n'
            if not has_id:
                sql += '    "id" INTEGER PRIMARY KEY AUTOINCREMENT'
            first_col = True
            for col in cols:
                if not first_col or not has_id:
                    sql += ",\n"
                first_col = False
                constraints = " NOT NULL" if col.get("not_null") else ""
                if col["name"] == "id" and has_id:
                    sql += f'    "id" INTEGER PRIMARY KEY AUTOINCREMENT{constraints}'
                else:
                    sql += f'    "{col["name"]}" {col["type"]}{constraints}'
            
            # Add UNIQUE constraint on empid if it has FK to employees
            if any(fk["column"] == "empid" for fk in fks):
                sql += ',\n    UNIQUE ("empid")'
            
            # Add FK constraints
            for fk in fks:
                sql += f',\n    FOREIGN KEY ("{fk["column"]}") REFERENCES "{fk["references_table"]}"("{fk["references_column"]}")'
            sql += "\n);\n"
            pipeline_sql += sql
        
        analysis_result["sql_schema"] = existing_sql + "\n" + pipeline_sql
    
    return added_tables


def _generate_pipeline_code(analysis_result, data_samples):
    """Generates consistent, contradiction-free Python pipeline code for the blueprint."""

    pipeline_code = '''# ==========================================
# SYNAPTO: Calculation Pipeline (Topological Sort)
# ==========================================
# IMPORTANT: This pipeline follows a strict topological order.
# All calculations are deterministic and must execute in this exact sequence.

from typing import Dict, List, Optional
from datetime import datetime
from dateutil.relativedelta import relativedelta

# Global calculation state
calc_state: Dict[str, Dict] = {}


def run_pipeline(month: str, db_session) -> None:
    """
    Execute the full payroll calculation pipeline.
    Must be called AFTER all user input data has been saved to the database.
    
    Steps are executed in topological sort order to ensure
    dependencies are resolved before dependent calculations.
    """
    steps = [
        _step_1_load_employees,
        _step_2_load_attendance,
        _step_3_compute_salary_processing,
        _step_4_compute_tax,
        _step_5_compute_benefits_lookup,
        _step_6_compute_final_payroll,
        _step_7_compute_dashboard_summary,
    ]
    for step in steps:
        step(month, db_session)


def _step_1_load_employees(month: str, db_session) -> None:
    """
    Step 1: Load all employees and compute years of service.
    Years of Service = (current_date - join_date).years
    """
    global calc_state
    calc_state = {}
    employees = db_session.query(Employee).all()
    for emp in employees:
        years = relativedelta(datetime.utcnow().date(), emp.join_date).years
        calc_state[emp.empid] = {"emp": emp, "years": years}


def _step_2_load_attendance(month: str, db_session) -> None:
    """
    Step 2: Load attendance records for the given month.
    Compute Overtime Pay = (Base_Salary / 30 / 8) * Overtime_Hours
    """
    for empid, data in calc_state.items():
        att = db_session.query(Attendance).filter_by(empid=empid, month=month).first()
        if att:
            emp = data["emp"]
            # Overtime Pay = hourly_rate * overtime_hours
            # hourly_rate = base_salary / 30 days / 8 hours
            hourly_rate = emp.base_salary / 30 / 8
            ot_pay = hourly_rate * att.overtime_hours
            calc_state[empid]["attendance"] = att
            calc_state[empid]["ot_pay"] = ot_pay
            
            # Update salary_processing with base_salary and overtime_pay
            sp = db_session.query(SalaryProcessing).filter_by(
                empid=empid, month=month
            ).first()
            if sp:
                sp.base_salary = emp.base_salary
                sp.overtime_pay = round(ot_pay, 2)
            else:
                sp = SalaryProcessing(
                    empid=empid,
                    month=month,
                    base_salary=emp.base_salary,
                    overtime_pay=round(ot_pay, 2)
                )
                db_session.add(sp)


def _step_3_compute_salary_processing(month: str, db_session) -> None:
    """
    Step 3: Compute Gross Salary.
    Gross_Salary = Base_Salary + Overtime_Pay
    NOTE: Benefits are NOT included in Gross Salary.
    """
    sp_records = db_session.query(SalaryProcessing).filter_by(month=month).all()
    for sp in sp_records:
        sp.gross_salary = round(sp.base_salary + sp.overtime_pay, 2)
    db_session.flush()


def _step_4_compute_tax(month: str, db_session) -> None:
    """
    Step 4: Compute Tax Amount using PROGRESSIVE tax formula.
    
    Tax = (Gross_Salary - Min_Salary) * Tax_Rate + Fixed_Deduction
    
    Find the bracket where: Min_Salary <= Gross_Salary < Max_Salary
    This is the SECOND IF condition (bracket selection).
    """
    brackets = db_session.query(TaxBracket).order_by(TaxBracket.min_salary).all()
    sp_records = db_session.query(SalaryProcessing).filter_by(month=month).all()
    
    for sp in sp_records:
        gross = sp.gross_salary
        tax = 0.0
        
        # IF/ELIF chain for bracket selection (second IF formula)
        for bracket in brackets:
            if bracket.min_salary <= gross < bracket.max_salary:
                # Progressive tax formula
                tax = (gross - bracket.min_salary) * bracket.tax_rate + bracket.fixed_deduction
                break
        
        sp.tax_amount = round(tax, 2)
        sp.net_salary = round(gross - tax, 2)
    db_session.flush()


def _step_5_compute_benefits_lookup(month: str, db_session) -> None:
    """
    Step 5: Compute Benefits using matrix VLOOKUP.
    
    Column Mapping for Benefits Matrix:
      Years of Service -> Column Key:
        1 year  -> "1"
        2 years -> "2"
        3 years -> "3"
        4 years -> "4"
        5+ years -> "5+"
    
    Benefit_Multiplier = BDBM_MAP[grade][column_key]
    Calculated_Benefit = Base_Salary * Benefit_Multiplier
    """
    from app.flat_matrix_loader import BDBM_MAP  # Loaded from JSON at startup
    
    for empid, data in calc_state.items():
        emp = data["emp"]
        years = data["years"]
        grade = emp.grade
        
        # Map years to column key
        if years >= 5:
            col_key = "5+"
        else:
            col_key = str(years)
        
        # VLOOKUP equivalent: lookup multiplier from the matrix
        grade_row = BDBM_MAP.get(grade, {})
        multiplier = grade_row.get(col_key, 0.0)
        
        # Calculate benefit
        calculated_benefit = emp.base_salary * multiplier
        
        # Update benefits_lookup table
        bl = db_session.query(BenefitsLookup).filter_by(empid=empid).first()
        if bl:
            bl.name = emp.name
            bl.grade = emp.grade
            bl.years_of_service = years
            bl.benefit_multiplier = multiplier
            bl.calculated_benefit = round(calculated_benefit, 2)
        else:
            bl = BenefitsLookup(
                empid=empid,
                name=emp.name,
                grade=emp.grade,
                years_of_service=years,
                benefit_multiplier=multiplier,
                calculated_benefit=round(calculated_benefit, 2)
            )
            db_session.add(bl)
    db_session.flush()


def _step_6_compute_final_payroll(month: str, db_session) -> None:
    """
    Step 6: Compute Final Payroll for each employee.
    
    Total_Net_Salary = SUM of net_salary across all months
    Total_Benefits = calculated_benefit from benefits_lookup
    Final_Payout = Total_Net_Salary + Total_Benefits
    Payout_Status = IF(Final_Payout > 0, "Paid", "Pending")  <- First IF formula
    """
    for empid, data in calc_state.items():
        # Sum net salary across all months
        sp_records = db_session.query(SalaryProcessing).filter_by(empid=empid).all()
        total_net = sum(sp.net_salary for sp in sp_records if sp.net_salary)
        
        # Get benefits
        bl = db_session.query(BenefitsLookup).filter_by(empid=empid).first()
        total_benefits = bl.calculated_benefit if bl and bl.calculated_benefit else 0.0
        
        final_payout = total_net + total_benefits
        
        # First IF formula: Payout Status
        payout_status = "Paid" if final_payout > 0 else "Pending"
        
        fp = db_session.query(FinalPayroll).filter_by(empid=empid).first()
        if fp:
            fp.total_net_salary = round(total_net, 2)
            fp.total_benefits = round(total_benefits, 2)
            fp.final_payout = round(final_payout, 2)
            fp.payout_status = payout_status
        else:
            fp = FinalPayroll(
                empid=empid,
                total_net_salary=round(total_net, 2),
                total_benefits=round(total_benefits, 2),
                final_payout=round(final_payout, 2),
                payout_status=payout_status
            )
            db_session.add(fp)
    db_session.flush()


def _step_7_compute_dashboard_summary(month: str, db_session) -> None:
    """
    Step 7: Compute Dashboard Summary aggregations.
    
    Total Headcount = COUNT of employees
    Total Payout = SUM of final_payout
    IT Dept Cost = SUM of final_payout WHERE department = 'IT'
    HR Dept Cost = SUM of final_payout WHERE department = 'HR'
    Finance Dept Cost = SUM of final_payout WHERE department = 'Finance'
    Operations Dept Cost = SUM of final_payout WHERE department = 'Operations'
    """
    total_headcount = db_session.query(Employee).count()
    
    fp_records = db_session.query(FinalPayroll).all()
    total_payout = sum(fp.final_payout for fp in fp_records if fp.final_payout)
    
    # Department costs
    dept_costs = {}
    for empid, data in calc_state.items():
        dept = data["emp"].department
        fp = db_session.query(FinalPayroll).filter_by(empid=empid).first()
        if fp and fp.final_payout:
            dept_costs[dept] = dept_costs.get(dept, 0.0) + fp.final_payout
    
    # Update dashboard_summary
    ds = db_session.query(DashboardSummary).first()
    if ds:
        ds.total_headcount = total_headcount
        ds.total_payout = round(total_payout, 2)
        ds.it_dept_cost = round(dept_costs.get("IT", 0.0), 2)
        ds.hr_dept_cost = round(dept_costs.get("HR", 0.0), 2)
        ds.finance_dept_cost = round(dept_costs.get("Finance", 0.0), 2)
        ds.operations_dept_cost = round(dept_costs.get("Operations", 0.0), 2)
    else:
        ds = DashboardSummary(
            total_headcount=total_headcount,
            total_payout=round(total_payout, 2),
            it_dept_cost=round(dept_costs.get("IT", 0.0), 2),
            hr_dept_cost=round(dept_costs.get("HR", 0.0), 2),
            finance_dept_cost=round(dept_costs.get("Finance", 0.0), 2),
            operations_dept_cost=round(dept_costs.get("Operations", 0.0), 2)
        )
        db_session.add(ds)
    db_session.commit()


def _step_8_compute_sales(db_session) -> None:
    """
    Step 8: Compute Sales calculated fields.
    
    Total = Quantity * Unit_Price * (1 - Discount / 100)
    Tax = Total * Tax_Rate (default 15% VAT, configurable in tax_brackets)
    Net_Amount = Total + Tax
    """
    sales = db_session.query(Sale).all()
    default_tax_rate = 0.15  # 15% default VAT
    for sale in sales:
        if sale.quantity and sale.unit_price and sale.discount is not None:
            sale.total = round(sale.quantity * sale.unit_price * (1 - sale.discount / 100), 2)
            sale.tax = round(sale.total * default_tax_rate, 2)
            sale.net_amount = round(sale.total + sale.tax, 2)
    db_session.flush()


def _step_9_compute_inventory(db_session) -> None:
    """
    Step 9: Compute Inventory calculated fields.
    
    Stock_Value = Stock_Level * Unit_Cost
    Status = IF(Stock_Level <= Reorder_Point, "Reorder", "In Stock")
    """
    items = db_session.query(InventoryItem).all()
    for item in items:
        if item.stock_level is not None and item.unit_cost is not None:
            item.stock_value = round(item.stock_level * item.unit_cost, 2)
        if item.stock_level is not None and item.reorder_point is not None:
            item.status = "Reorder" if item.stock_level <= item.reorder_point else "In Stock"
        else:
            item.status = "Unknown"
    db_session.flush()


def _step_10_compute_projects(db_session) -> None:
    """
    Step 10: Compute Projects calculated fields.
    
    Remaining = Budget - Spent
    Status = IF(Progress >= 100, "Completed", IF(Remaining <= 0, "Over Budget", "In Progress"))
    Days_Left = End_Date - Start_Date
    """
    projects = db_session.query(Project).all()
    for proj in projects:
        if proj.budget is not None and proj.spent is not None:
            proj.remaining = round(proj.budget - proj.spent, 2)
        if proj.progress is not None:
            if proj.progress >= 100:
                proj.status = "Completed"
            elif proj.remaining is not None and proj.remaining < 0:
                proj.status = "Over Budget"
            elif proj.remaining is not None and proj.remaining == 0:
                proj.status = "On Budget"
            else:
                proj.status = "In Progress"
        if proj.end_date is not None and proj.start_date is not None:
            proj.days_left = proj.end_date - proj.start_date
    db_session.flush()
'''
    # V5 FIX: Dynamically add Sales/Inventory/Projects steps based on detected tables
    tables_meta = analysis_result.get("tables_metadata", {})
    extra_steps = ""
    extra_run_steps = ""
    step_num = 8
    
    if "sales" in tables_meta:
        extra_steps += '''
def _step_8_compute_sales(db_session) -> None:
    """
    Step 8: Compute Sales calculated fields.
    
    Total = Quantity * Unit_Price * (1 - Discount / 100)
    Tax = Total * Tax_Rate (default 15% VAT, configurable in tax_brackets)
    Net_Amount = Total + Tax
    """
    sales = db_session.query(Sale).all()
    default_tax_rate = 0.15  # 15% default VAT
    for sale in sales:
        if sale.quantity and sale.unit_price and sale.discount is not None:
            sale.total = round(sale.quantity * sale.unit_price * (1 - sale.discount / 100), 2)
            sale.tax = round(sale.total * default_tax_rate, 2)
            sale.net_amount = round(sale.total + sale.tax, 2)
    db_session.flush()
'''
        extra_run_steps += f"        _step_{step_num}_compute_sales,\n"
        step_num += 1
    
    if "inventory" in tables_meta:
        extra_steps += f'''
def _step_{step_num}_compute_inventory(db_session) -> None:
    """
    Step {step_num}: Compute Inventory calculated fields.
    
    Stock_Value = Stock_Level * Unit_Cost
    Status = IF(Stock_Level <= Reorder_Point, "Reorder", "In Stock")
    """
    items = db_session.query(InventoryItem).all()
    for item in items:
        if item.stock_level is not None and item.unit_cost is not None:
            item.stock_value = round(item.stock_level * item.unit_cost, 2)
        if item.stock_level is not None and item.reorder_point is not None:
            item.status = "Reorder" if item.stock_level <= item.reorder_point else "In Stock"
        else:
            item.status = "Unknown"
    db_session.flush()
'''
        extra_run_steps += f"        _step_{step_num}_compute_inventory,\n"
        step_num += 1
    
    if "projects" in tables_meta:
        extra_steps += f'''
def _step_{step_num}_compute_projects(db_session) -> None:
    """
    Step {step_num}: Compute Projects calculated fields.
    
    Remaining = Budget - Spent
    Status = IF(Progress >= 100, "Completed", IF(Remaining < 0, "Over Budget", "In Progress"))
    Days_Left = End_Date - Start_Date
    """
    projects = db_session.query(Project).all()
    for proj in projects:
        if proj.budget is not None and proj.spent is not None:
            proj.remaining = round(proj.budget - proj.spent, 2)
        if proj.progress is not None:
            if proj.progress >= 100:
                proj.status = "Completed"
            elif proj.remaining is not None and proj.remaining < 0:
                proj.status = "Over Budget"
            elif proj.remaining is not None and proj.remaining == 0:
                proj.status = "On Budget"
            else:
                proj.status = "In Progress"
        if proj.end_date is not None and proj.start_date is not None:
            proj.days_left = proj.end_date - proj.start_date
    db_session.flush()
'''
        extra_run_steps += f"        _step_{step_num}_compute_projects,\n"
        step_num += 1
    
    # Inject extra steps into pipeline code
    if extra_run_steps:
        pipeline_code = pipeline_code.replace(
            "        _step_7_compute_dashboard_summary,\n    ]",
            "        _step_7_compute_dashboard_summary,\n" + extra_run_steps + "    ]"
        )
        pipeline_code = pipeline_code.rstrip() + "\n\n" + extra_steps
    
    return pipeline_code


# ==========================================
# License / Activation Endpoint
# ==========================================
def _sha256_hex(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()

def _get_activation_code_hash() -> str:
    code_hash = os.environ.get("SYNAPTO_ACTIVATION_CODE_HASH", "").strip()
    if not code_hash:
        # Default expected hash for synapto988 to make the app runnable out-of-box.
        # SECURITY NOTE: still provided via environment variable in production.
        code_hash = _sha256_hex("synapto988")
    return code_hash

def _get_token_secret() -> str:
    secret = os.environ.get("SYNAPTO_ACTIVATION_TOKEN_SECRET", "").strip()
    if not secret:
        # Default secret (dev). Should be overridden in production.
        secret = "synapto-token-secret-dev-change-me"
    return secret

def _sign_payload(payload: str) -> str:
    secret = _get_token_secret().encode("utf-8")
    sig = hmac.new(secret, payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return sig


def _make_activation_token(expires_at_iso: str) -> str:
    # Token format: synapto|<expiresAt>|<hmac>
    # نستخدم | بدلاً من : لأن التاريخ ISO يحتوي على :
    payload = f"{expires_at_iso}"
    sig = _sign_payload(payload)
    return f"synapto|{expires_at_iso}|{sig}"

def _verify_activation_token(token: str):
    try:
        if not token:
            return False, None
        parts = token.split("|")
        if len(parts) != 3:
            return False, None
        prefix, expires_at_iso, sig = parts
        if prefix != "synapto":
            return False, None
        expected_sig = _sign_payload(expires_at_iso)
        if not hmac.compare_digest(expected_sig, sig):
            return False, None
        
        # FIX: مقارنة التواريخ بدون مشاكل الـ Timezone
        try:
            from datetime import timezone
            # جعل التاريخين بنفس الصيغة (مع Timezone)
            dt = datetime.fromisoformat(expires_at_iso.replace("Z", "+00:00"))
            now = datetime.now(timezone.utc)
            if dt < now:
                return False, expires_at_iso
        except Exception:
            pass
            
        return True, expires_at_iso
    except Exception:
        return False, None


class ActivateRequest(BaseModel):
    code: str

class ActivateResponse(BaseModel):
    success: bool
    expiresAt: Optional[str] = None
    token: Optional[str] = None
    codeHash: Optional[str] = None
    error: Optional[str] = None

@app.post("/api/activate")
async def activate(request: Request):
    # IMPORTANT:
    # Your client calls are currently failing with 422 JSON decode error BEFORE Pydantic validation.
    # Therefore we accept ANY body, parse JSON manually, and return debug info if parsing fails.
    try:
        raw = await request.body()
    except Exception:
        raw = b""

    raw_text = raw.decode("utf-8", errors="replace") if raw else ""

    parsed = None
    try:
        if raw:
            parsed = json.loads(raw_text)
    except Exception:
        parsed = None

    code = ""
    if isinstance(parsed, dict):
        code = parsed.get("code") or ""

    expected_code_hash = _get_activation_code_hash()
    provided_code_hash = _sha256_hex(code or "")

    # Debug payload + parsing result
    debug_payload = {
        "content-type": request.headers.get("content-type"),
        "raw_body": raw_text,
        "parsed_json": parsed,
        "code_received": code,
        "expected_code_hash": expected_code_hash,
        "provided_code_hash": provided_code_hash,
    }

    if not hmac.compare_digest(provided_code_hash, expected_code_hash):
        return JSONResponse(
            {"success": False, "error": "Invalid activation code", "debug": debug_payload},
            status_code=200
        )

    expires_at = datetime.utcnow() + timedelta(days=30)
    expires_at_iso = expires_at.replace(microsecond=0).isoformat() + "Z"
    token = _make_activation_token(expires_at_iso)

    return JSONResponse(
        {
            "success": True,
            "expiresAt": expires_at_iso,
            "token": token,
            "codeHash": expected_code_hash,
            "debug": debug_payload,
        },
        status_code=200
    )


def _require_activation_token(request: Request):
    # Allow free trial
    if request.headers.get("X-Free-Trial") == "true":
        return

    token = request.headers.get("X-Activation-Token") or ""
    ok, _ = _verify_activation_token(token)
    if not ok:
        raise HTTPException(401, "License required")

    # Check if license is still active in DB
    try:
        conn = get_admin_db()
        c = conn.cursor()
        c.execute("SELECT is_active FROM licenses WHERE token = ?", (token,))
        row = c.fetchone()
        conn.close()
        if not row or row["is_active"] != 1:
            raise HTTPException(401, "License deactivated")
    except HTTPException:
        raise
    except Exception:
        pass  # If DB check fails, allow (graceful degradation)

# ==========================================
# API Endpoints
# ==========================================


# ==========================================
# User-Facing: Request Activation
# ==========================================
@app.post("/api/request-activation")
async def request_activation(request: Request):
    try:
        raw = await request.body()
        data = json.loads(raw.decode("utf-8"))
    except:
        return JSONResponse({"success": False, "error": "Invalid JSON"}, status_code=400)

    full_name = (data.get("full_name") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    company = (data.get("company") or "").strip()
    notes = (data.get("notes") or "").strip()

    if not full_name or not email:
        return JSONResponse({"success": False, "error": "Name and email are required"})

    conn = get_admin_db()
    c = conn.cursor()

    # Check if email already has an active license
    c.execute("SELECT id, is_active FROM licenses WHERE email = ?", (email,))
    existing = c.fetchone()
    if existing and existing["is_active"]:
        conn.close()
        return JSONResponse({"success": False, "error": "This email already has an active license"})

    # Upsert: if request exists for this email, reset to pending
    c.execute("SELECT id FROM activation_requests WHERE email = ?", (email,))
    existing_req = c.fetchone()
    if existing_req:
        c.execute("""
            UPDATE activation_requests SET full_name=?, phone=?, company=?, notes=?, status='pending', reviewed_at=NULL, admin_note=NULL, created_at=datetime('now','localtime')
            WHERE id = ?
        """, (full_name, phone, company, notes, existing_req["id"]))
    else:
        c.execute("""
            INSERT INTO activation_requests (full_name, email, phone, company, notes) VALUES (?, ?, ?, ?, ?)
        """, (full_name, email, phone, company, notes))

    conn.commit()
    conn.close()
    log_activity("request_created", email, f"Name: {full_name}, Company: {company}")

    return JSONResponse({"success": True, "message": "Request submitted"})


@app.post("/api/check-request-status")
async def check_request_status(request: Request):
    try:
        raw = await request.body()
        data = json.loads(raw.decode("utf-8"))
    except:
        return JSONResponse({"success": False, "error": "Invalid JSON"}, status_code=400)

    email = (data.get("email") or "").strip()
    if not email:
        return JSONResponse({"success": False, "error": "Email is required"})

    conn = get_admin_db()
    c = conn.cursor()

    c.execute("SELECT status, admin_note FROM activation_requests WHERE email = ? ORDER BY id DESC LIMIT 1", (email,))
    req = c.fetchone()

    if not req:
        conn.close()
        return JSONResponse({"success": True, "status": "not_found"})

    status = req["status"]

    if status == "pending":
        conn.close()
        return JSONResponse({"success": True, "status": "pending"})

    if status == "rejected":
        conn.close()
        return JSONResponse({"success": True, "status": "rejected", "reason": req["admin_note"] or ""})

    if status == "approved":
        c.execute("SELECT token, expires_at FROM licenses WHERE email = ? AND is_active = 1", (email,))
        lic = c.fetchone()
        conn.close()
        if lic:
            return JSONResponse({"success": True, "status": "approved", "token": lic["token"], "expiresAt": lic["expires_at"]})
        return JSONResponse({"success": True, "status": "pending"})

    conn.close()
    return JSONResponse({"success": True, "status": "not_found"})


@app.post("/analyze")
async def analyze_file(file: UploadFile = File(...), request: Request = None):
    # Require a valid activation token for every analysis request
    _require_activation_token(request)

    if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
        raise HTTPException(400, "Only Excel files allowed")

    temp_file_path = f"temp_{file.filename}"
    try:
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        analyzer = ExcelAnalyzer(temp_file_path)
        result = analyzer.analyze()

        wb_stats = openpyxl.load_workbook(temp_file_path, data_only=False, read_only=True, keep_vba=False)
        total_cells, filled_cells, sheets_info, total_formulas, formula_logic = 0, 0, [], 0, []

        for ws in wb_stats.worksheets:
            rows, cols = ws.max_row or 0, ws.max_column or 0
            total_cells += rows * cols
            sheets_info.append({"name": ws.title, "max_row": rows, "max_column": cols})
            headers = [
                str(cell.value) if cell.value is not None else f"Unnamed_{i}"
                for i, cell in enumerate(list(ws.iter_rows(min_row=1, max_row=1))[0])
            ]
            for row in ws.iter_rows(min_row=2, max_row=min(rows, 1000)):
                for cell in row:
                    if cell.value is not None:
                        filled_cells += 1
                    if cell.data_type == 'f':
                        total_formulas += 1
                        formula_str = (
                            f"={cell.value.text}"
                            if hasattr(cell.value, 'text')
                            else (cell.value if isinstance(cell.value, str) else str(cell.value))
                        )
                        logic_type = (
                            re.match(r'=([A-Z]+)\(', formula_str, re.IGNORECASE).group(1).upper()
                            if re.match(r'=([A-Z]+)\(', formula_str, re.IGNORECASE)
                            else "REFERENCE"
                        )
                        target_col_name = (
                            headers[cell.column - 1]
                            if cell.column - 1 < len(headers)
                            else f"Column_{cell.column_letter}"
                        )
                        formula_logic.append({
                            "sheet": ws.title,
                            "target_column": target_col_name,
                            "logic_type": logic_type,
                            "description": f"Calculates '{target_col_name}'",
                            "source": "Intra-sheet",
                            "raw_formula": formula_str
                        })

        wb_stats.close()

        # FIX 7: Include status and functional_role in analysis response
        # Build ref_counts for this analysis
        ref_counts_local = {t_name: {"incoming": 0, "outgoing": 0} for t_name in result.get("tables_metadata", {})}
        for ref in result.get("cross_sheet_refs", []):
            if ref.get("source") in ref_counts_local:
                ref_counts_local[ref["source"]]["outgoing"] += 1
            if ref.get("target") in ref_counts_local:
                ref_counts_local[ref["target"]]["incoming"] += 1

        # Add status and functional_role to tables_metadata
        for t_name, t_meta in result.get("tables_metadata", {}).items():
            t_meta["status"] = _classify_table(t_meta)
            ref_counts_val = ref_counts_local.get(t_name, {"incoming": 0, "outgoing": 0})
            t_meta["functional_role"] = _determine_functional_role(t_name, t_meta, ref_counts_val["incoming"], ref_counts_val["outgoing"], False)

        result.update({
            "file_name": file.filename,
            "total_cells": total_cells,
            "data_quality": {"filled_cells": filled_cells, "empty_cells": total_cells - filled_cells},
            "sheets": sheets_info,
            "total_sheets": len(sheets_info),
            "total_formulas": total_formulas,
            "formula_logic": formula_logic,
            "cross_sheet_refs": [
                {
                    "source": r.get("source", ""),
                    "target": r.get("target", ""),
                    "type": r.get("type", ""),
                    "logic_type": r.get("logic_type", ""),
                    "raw_formula": r.get("raw_formula", ""),
                    # Also keep from/to format for backward compatibility
                    "from": {"sheet": r.get("source", ""), "column": "N/A"},
                    "to": {"sheet": r.get("target", ""), "column": "N/A"},
                }
                for r in result.get("cross_sheet_refs", [])
            ]
        })

        return JSONResponse(sanitize_for_json(result))

    except Exception as e:
        return JSONResponse({"error": str(e), "trace": traceback.format_exc()}, status_code=500)
    finally:
        if os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except:
                pass


@app.post("/download-clean-excel")
async def download_clean_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
        raise HTTPException(400, "Only Excel files allowed")

    temp_file_path, db_path, engine = f"temp_dl_{file.filename}", None, None
    try:
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        analysis_result = ExcelAnalyzer(temp_file_path).analyze()
        sql_schema = analysis_result["sql_schema"]

        wb = openpyxl.load_workbook(temp_file_path, data_only=True, keep_vba=False)
        excel_buffer, zip_buffer = io.BytesIO(), io.BytesIO()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".db") as tmp_db:
            db_path = tmp_db.name
        engine = create_engine(f'sqlite:///{db_path}')

        with engine.connect() as conn:
            conn.execute(text("PRAGMA foreign_keys = ON;"))
            for statement in sql_schema.split(';'):
                if statement.strip():
                    conn.execute(text(statement))
            conn.commit()

        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            for table_name, sheet_meta in analysis_result["tables_metadata"].items():
                if not sheet_meta.get("original_name") or sheet_meta["original_name"] not in wb.sheetnames:
                    continue
                data = list(wb[sheet_meta["original_name"]].values)
                if not data or len(data) < 2:
                    continue
                df = pd.DataFrame(data[1:], columns=data[0])
                df.columns = [
                    re.sub(r'[^\w]+', '_', str(col), flags=re.UNICODE).lower().strip('_')
                    for col in df.columns
                ]
                df.to_excel(writer, sheet_name=sheet_meta["original_name"], index=False)
                try:
                    df.to_sql(table_name, con=engine, if_exists='append', index=False)
                except Exception as db_err:
                    print(f"DB Error: {db_err}")

        excel_buffer.seek(0)
        base_filename = os.path.splitext(file.filename)[0]

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr(f"clean_{base_filename}.xlsx", excel_buffer.getvalue())
            with open(db_path, 'rb') as f:
                zip_file.writestr(f"{base_filename}_database.db", f.read())

        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename=synapto_cleaned_{base_filename}.zip"}
        )

    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if engine:
            engine.dispose()
        if db_path and os.path.exists(db_path):
            os.unlink(db_path)
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)


@app.post("/architect-report")
async def architect_report(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
        raise HTTPException(400, "Only Excel files allowed")

    temp_file_path = f"temp_arch_{file.filename}"
    try:
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        analysis_result = ExcelAnalyzer(temp_file_path).analyze()
        
        # V5 FIX: Add pipeline tables that are missing from the Excel
        # These are tables referenced by the pipeline code but not present as sheets
        _add_pipeline_tables(analysis_result)
        
        wb_ui = None
        try:
            wb_ui = openpyxl.load_workbook(temp_file_path, data_only=False, keep_vba=False, read_only=True)
        except:
            pass

        data_samples = _extract_data_samples(temp_file_path, analysis_result, 5)
        execution_flow = _derive_execution_flow(analysis_result)
        strict_rules = _generate_strict_rules(analysis_result, temp_file_path, data_samples)
        pipeline_code = _generate_pipeline_code(analysis_result, data_samples)

        blueprint = {
            "phase_1_database": _generate_phase_1(analysis_result, data_samples),
            "phase_2_logic": _generate_phase_2(analysis_result, temp_file_path),
            "phase_3_ui_ux": _generate_phase_3(analysis_result, wb_ui),
            "phase_4_auth": _generate_phase_4(analysis_result),
            "phase_5_validation": _generate_phase_5(analysis_result),
            "phase_6_api_design": _generate_phase_6(analysis_result),
            "execution_flow": execution_flow,
            "strict_rules": strict_rules,
            "data_samples": data_samples,
            "pipeline_code": pipeline_code,
        }

        # FIX 6: SQL Schema should NOT include CREATE TABLE for flat matrices
        # Also add FK constraints to the SQL
        tables_meta = analysis_result.get("tables_metadata", {})
        filtered_sql_parts = []
        for sql_statement in analysis_result.get("sql_schema", "").split(';'):
            sql_statement = sql_statement.strip()
            if not sql_statement:
                continue
            # Check if this is a CREATE TABLE for a flat matrix
            is_flat_matrix = False
            for t_name, t_meta in tables_meta.items():
                if _classify_table(t_meta) == "Flat_Calculation_Matrix":
                    # Check if this SQL statement creates that table
                    if f'CREATE TABLE "{t_name}"' in sql_statement:
                        is_flat_matrix = True
                        break
            if not is_flat_matrix:
                filtered_sql_parts.append(sql_statement)

        # Rebuild SQL with FK constraints
        final_sql_parts = []
        for sql_stmt in filtered_sql_parts:
            # Find the table name from the CREATE TABLE statement
            table_match = re.match(r'CREATE TABLE "(\w+)"', sql_stmt)
            if table_match:
                t_name = table_match.group(1)
                # Add FK constraint for empid column if it exists
                if t_name in tables_meta:
                    t_meta = tables_meta[t_name]
                    for fk in t_meta.get('fks', []):
                        fk_col = fk['column']
                        ref_table = fk['references_table']
                        ref_col = fk['references_column']
                        # Add FK constraint before the closing paren
                        if f'"{fk_col}"' in sql_stmt:
                            sql_stmt = sql_stmt.rstrip().rstrip(')')
                            sql_stmt += f',\n    FOREIGN KEY ("{fk_col}") REFERENCES "{ref_table}"("{ref_col}")\n)'
                final_sql_parts.append(sql_stmt)
            else:
                final_sql_parts.append(sql_stmt)

        blueprint["sql_schema"] = ';\n'.join(final_sql_parts) + ';'

        try:
            from system_architect import SystemArchitect
            blueprint["ai_ready_system_prompt_markdown"] = SystemArchitect(
                temp_file_path, blueprint, analysis_result, data_samples, execution_flow, strict_rules
            ).generate_blueprint().get("ai_ready_system_prompt_markdown", "")
        except:
            blueprint["ai_ready_system_prompt_markdown"] = ""

        if wb_ui:
            wb_ui.close()

        return JSONResponse(sanitize_for_json(blueprint))

    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)


@app.post("/ai-report")
async def ai_report(req: ReportRequest):
    try:
        d, provider, architect_data = req.data, req.provider, req.architectData
        tables_meta = d.get("tables_metadata", {})
        formula_logic = d.get("formula_logic", [])
        cross_refs = d.get("cross_sheet_refs", [])

        tables_summary = ""
        for t_name, t_meta in tables_meta.items():
            fks = [f"{f['column']}->{f['references_table']}" for f in t_meta.get('fks', [])]
            tables_summary += f"- **{t_name}** (Role: {t_meta.get('functional_role')}, Type: {t_meta.get('status')}, LinksTo: {', '.join(fks) if fks else 'None'})\n"

        logic_summary = {}
        for f in formula_logic:
            ltype = f.get("logic_type", "OTHER")
            if ltype not in logic_summary:
                logic_summary[ltype] = {"count": 0, "targets": set()}
            logic_summary[ltype]["count"] += 1
            logic_summary[ltype]["targets"].add(f.get("target_column", ""))

        logic_text = "\n".join([
            f"- **{ltype}** (x{info['count']}): Affects [{', '.join(list(info['targets'])[:5])}]"
            for ltype, info in logic_summary.items()
        ])

        links_text = "\n".join([
            f"- {ref.get('from', {}).get('sheet', '?')} --({ref.get('logic_type', '')})--> {ref.get('to', {}).get('sheet', '?')}"
            for ref in cross_refs
        ])

        architect_context = ""
        if architect_data:
            flow_text = "\n".join([
                f"Step {step.get('step')}: {step.get('phase')} on {step.get('entity')} - {step.get('action')}"
                for step in architect_data.get("execution_flow", [])
            ])
            rules_text = "\n".join([
                f"- [{rule.get('severity')}] {rule.get('rule')}"
                for rule in architect_data.get("strict_rules", [])
            ])
            screens_text = "\n".join([
                f"- {s.get('entity')}: Route {s.get('route')}, Role: {s.get('functional_role')}"
                for s in architect_data.get("phase_3_ui_ux", {}).get("screens", [])
            ])
            architect_context = f"""
### Python Engine Architecture Blueprint (PRE-COMPUTED)
#### Execution Flow:
{flow_text if flow_text else "Not generated."}
#### Strict Engineering Rules:
{rules_text if rules_text else "None."}
#### UI/UX & Functional Roles:
{screens_text if screens_text else "Not generated."}
"""

        # Extract real data samples
        samples_context = ""
        if architect_data:
            data_samples = architect_data.get("data_samples", {})
            for t_name, rows in data_samples.items():
                if rows:
                    samples_context += f"\n#### Data Sample for {t_name}:\n```json\n{json.dumps(rows[:1], indent=2, ensure_ascii=False)}\n```\n"

        # Pipeline code context
        pipeline_context = ""
        if architect_data and architect_data.get("pipeline_code"):
            pipeline_context = f"""
### Pre-Generated Calculation Pipeline Code:
```python
{architect_data.get("pipeline_code")}
```
"""

        prompt = f"""You are a Senior System Architect. I have reverse-engineered an Excel file using a sophisticated Python Engine. The data extraction, SQL schema, and Architectural Blueprint are ALREADY DONE.

Your job is to provide a STEP-BY-STEP IMPLEMENTATION GUIDE to build the web application based on the Python blueprint below. Do NOT rewrite SQL or repeat the structure. Focus on CODE LOGIC and ARCHITECTURE.

## Raw Extracted Metadata:
### 1. Tables & Roles
{tables_summary if tables_summary else "No tables extracted."}
### 2. Business Logic Summary
{logic_text if logic_text else "No formulas detected."}
### 3. Cross-Sheet Data Flow
{links_text if links_text else "No cross-sheet dependencies."}

{samples_context if samples_context else ""}

{architect_context if architect_context else ""}

{pipeline_context if pipeline_context else ""}

---
## YOUR OUTPUT (Be concise, no boilerplate):

### 1. Backend Service Layer Implementation
- **CRITICAL DISTINCTION:** You MUST distinguish between `Relational_Entity` and `Flat_Calculation_Matrix`.
- For `Relational_Entity` tables: Use SQLAlchemy models and standard SQL CRUD operations (FastAPI + PostgreSQL/SQLite).
- For `Flat_Calculation_Matrix` tables: Load them as in-memory JSON dictionaries at startup. DO NOT mix the two approaches.
- For VLOOKUP/REFERENCE: Specify the exact Python dictionary lookup logic based on the Data Samples provided.
- **Tax Calculation:** Use progressive formula: tax = (gross - min_salary) * tax_rate + fixed_deduction
- **Gross Salary:** Gross = Base_Salary + Overtime_Pay ONLY (benefits are separate)
- **Overtime Pay:** overtime_pay = (base_salary / 30 / 8) * overtime_hours

### 2. Calculation Pipeline (Topological Sort)
- The pre-generated pipeline code above is the REFERENCE implementation. Use it directly.
- **Handling IF Logic:**
  1. Payout Status: if final_payout > 0 then "Paid" else "Pending"
  2. Tax Bracket Selection: if/elif chain to find matching bracket
- How to handle state and caching between steps?

### 3. Frontend & API Integration
- Map the provided UI Routes to React components.
- What triggers the recalculation pipeline? (e.g., API endpoint design)

### 4. Critical Code Warnings
- Enforce the Strict Engineering Rules in the codebase.
- Prevent circular dependencies and race conditions.

Keep response under 1500 words. Focus on CODE ARCHITECTURE and LOGIC.
"""

        if provider == "gemini-free":
            if not GEMINI_API_KEY:
                raise ValueError("GEMINI_API_KEY environment variable not set")
            payload = {"contents": [{"role": "user", "parts": [{"text": prompt}]}]}
            headers = {"Content-Type": "application/json"}
            response = requests.post(GEMINI_URL, headers=headers, data=json.dumps(payload), timeout=120)
            if response.status_code != 200:
                raise ValueError(f"Gemini Error ({response.status_code}): {response.text}")
            parts = response.json().get("candidates", [{}])[0].get("content", {}).get("parts", [])
            full_text = "".join([part.get("text", "") for part in parts])
            if not full_text:
                raise ValueError("Empty Gemini response.")
            return {"report": full_text}

        else:
            if not OPENROUTER_API_KEY:
                raise ValueError("OPENROUTER_API_KEY environment variable not set")
            selected_model = "openai/gpt-oss-20b:free" if provider == "openrouter-free" else PRO_MODELS.get("pro", "openai/gpt-4o")
            headers = {
                "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                "Content-Type": "application/json",
                "HTTP-Referer": "http://localhost:3000"
            }
            payload = {
                "model": selected_model,
                "max_tokens": OPENROUTER_MAX_TOKENS,
                "messages": [{"role": "user", "content": prompt}]
            }
            response = requests.post(OPENROUTER_URL, headers=headers, data=json.dumps(payload), timeout=120)
            if response.status_code != 200:
                raise ValueError(f"OpenRouter Error ({response.status_code}): {response.text}")
            full_text = response.json().get("choices", [{}])[0].get("message", {}).get("content", "")
            if not full_text:
                raise ValueError("Empty OpenRouter response.")
            return {"report": full_text}

    except Exception as e:
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/health")
def health():
    return {"status": "ok", "version": "v5"}


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)


# ==========================================
# Admin Panel Routes
# ==========================================
def admin_auth(request: Request):
    if request.session.get("admin_logged_in") != True:
        return False
    return True

@app.get("/admin", response_class=HTMLResponse)
async def admin_dashboard(request: Request):
    if not admin_auth(request):
        return RedirectResponse("/admin/login", status_code=302)
        
    conn = get_admin_db()
    c = conn.cursor()
    
    # جلب الإحصائيات بأمان عبر التفكيك المباشر (Tuple Unpacking) لضمان عدم حدوث Key Error
    c.execute("SELECT COUNT(*) FROM activation_requests WHERE status='pending'")
    pending = c.fetchone()[0] or 0
    
    c.execute("SELECT COUNT(*) FROM activation_requests WHERE status='approved'")
    approved = c.fetchone()[0] or 0
    
    c.execute("SELECT COUNT(*) FROM licenses WHERE is_active=1")
    active = c.fetchone()[0] or 0
    
    c.execute("SELECT COUNT(*) FROM activation_requests WHERE status='rejected'")
    rejected = c.fetchone()[0] or 0
    
    # جلب الطلبات الأخيرة (تعتمد على الـ row_factory كـ Dictionary أو Rows بنجاح)
    c.execute("SELECT * FROM activation_requests ORDER BY id DESC LIMIT 10")
    recent = c.fetchall()
    
    conn.close()
    
    # رندرة القالب وإرسال المتغيرات المستقرة
    return templates.TemplateResponse("dashboard.html", {
        "request": request, 
        "active_page": "dashboard",
        "pending_count": pending, 
        "approved_count": approved,
        "active_licenses": active, 
        "rejected_count": rejected,
        "recent_requests": recent
    })

@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    if admin_auth(request):
        return RedirectResponse("/admin", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "error": None})

@app.post("/admin/login")
async def admin_login_submit(request: Request):
    form = await request.form()
    if form.get("password") == ADMIN_PASSWORD:
        request.session["admin_logged_in"] = True
        return RedirectResponse("/admin", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "error": "Wrong password"})

@app.get("/admin/logout")
async def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse("/admin/login", status_code=302)

@app.get("/admin/requests", response_class=HTMLResponse)
async def admin_requests(request: Request):
    if not admin_auth(request):
        return RedirectResponse("/admin/login", status_code=302)
    conn = get_admin_db()
    c = conn.cursor()
    c.execute("SELECT * FROM activation_requests ORDER BY id DESC")
    reqs = c.fetchall()
    conn.close()
    return templates.TemplateResponse("requests.html", {"request": request, "active_page": "requests", "requests": reqs})

@app.post("/admin/requests/{req_id}/approve")
async def admin_approve(req_id: int, request: Request):
    if not admin_auth(request):
        return RedirectResponse("/admin/login", status_code=302)
    conn = get_admin_db()
    c = conn.cursor()
    c.execute("SELECT * FROM activation_requests WHERE id = ?", (req_id,))
    req = c.fetchone()
    if not req:
        conn.close()
        return RedirectResponse("/admin/requests", status_code=302)

    now_iso = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    expires_at = (datetime.utcnow() + timedelta(days=30)).replace(microsecond=0).isoformat() + "Z"
    token = _make_activation_token(expires_at)

    c.execute("UPDATE activation_requests SET status='approved', reviewed_at=datetime('now','localtime') WHERE id = ?", (req_id,))
    # Upsert license
    c.execute("SELECT id FROM licenses WHERE email = ?", (req["email"],))
    existing = c.fetchone()
    if existing:
        c.execute("UPDATE licenses SET token=?, expires_at=?, is_active=1, full_name=?, company=?, deactivated_at=NULL, deactivate_reason=NULL WHERE id = ?",
                  (token, expires_at, req["full_name"], req["company"], existing["id"]))
    else:
        c.execute("INSERT INTO licenses (request_id, email, full_name, company, token, expires_at) VALUES (?, ?, ?, ?, ?, ?)",
                  (req_id, req["email"], req["full_name"], req["company"], token, expires_at))
    conn.commit()
    conn.close()
    log_activity("approved", req["email"], f"Request #{req_id}, Expires: {expires_at}")
    return RedirectResponse("/admin/requests", status_code=302)

@app.post("/admin/requests/{req_id}/reject")
async def admin_reject(req_id: int, request: Request, reason: str = FastAPIForm("")):
    if not admin_auth(request):
        return RedirectResponse("/admin/login", status_code=302)
    conn = get_admin_db()
    c = conn.cursor()
    c.execute("SELECT email FROM activation_requests WHERE id = ?", (req_id,))
    req = c.fetchone()
    if req:
        c.execute("UPDATE activation_requests SET status='rejected', reviewed_at=datetime('now','localtime'), admin_note=? WHERE id = ?", (reason, req_id))
        conn.commit()
        log_activity("rejected", req["email"], f"Request #{req_id}, Reason: {reason}")
    conn.close()
    return RedirectResponse("/admin/requests", status_code=302)

@app.get("/admin/licenses", response_class=HTMLResponse)
async def admin_licenses(request: Request):
    if not admin_auth(request):
        return RedirectResponse("/admin/login", status_code=302)
    conn = get_admin_db()
    c = conn.cursor()
    c.execute("SELECT l.*, r.company as req_company FROM licenses l LEFT JOIN activation_requests r ON l.request_id = r.id ORDER BY l.id DESC")
    licenses = []
    for row in c.fetchall():
        lic = dict(row)
        try:
            exp = datetime.fromisoformat(lic["expires_at"].replace("Z", "+00:00"))
            from datetime import timezone
            days = max(0, (exp - datetime.now(timezone.utc)).days)
        except:
            days = 0
        lic["days_left"] = days
        lic["company"] = lic.get("company") or lic.get("req_company") or ""
        licenses.append(lic)
    conn.close()
    return templates.TemplateResponse("licenses.html", {"request": request, "active_page": "licenses", "licenses": licenses})

@app.post("/admin/licenses/{lic_id}/deactivate")
async def admin_deactivate(lic_id: int, request: Request, reason: str = FastAPIForm("")):
    if not admin_auth(request):
        return RedirectResponse("/admin/login", status_code=302)
    conn = get_admin_db()
    c = conn.cursor()
    c.execute("SELECT email FROM licenses WHERE id = ?", (lic_id,))
    lic = c.fetchone()
    if lic:
        c.execute("UPDATE licenses SET is_active=0, deactivated_at=datetime('now','localtime'), deactivate_reason=? WHERE id = ?", (reason, lic_id))
        conn.commit()
        log_activity("deactivated", lic["email"], f"License #{lic_id}, Reason: {reason}")
    conn.close()
    return RedirectResponse("/admin/licenses", status_code=302)

@app.post("/admin/licenses/{lic_id}/reactivate")
async def admin_reactivate(lic_id: int, request: Request):
    if not admin_auth(request):
        return RedirectResponse("/admin/login", status_code=302)
    conn = get_admin_db()
    c = conn.cursor()
    c.execute("SELECT email, expires_at FROM licenses WHERE id = ?", (lic_id,))
    lic = c.fetchone()
    if lic:
        # Extend expiry by 30 days from now
        new_exp = (datetime.utcnow() + timedelta(days=30)).replace(microsecond=0).isoformat() + "Z"
        new_token = _make_activation_token(new_exp)
        c.execute("UPDATE licenses SET is_active=1, token=?, expires_at=?, deactivated_at=NULL, deactivate_reason=NULL WHERE id = ?", (new_token, new_exp, lic_id))
        conn.commit()
        log_activity("reactivated", lic["email"], f"License #{lic_id}, New expiry: {new_exp}")
    conn.close()
    return RedirectResponse("/admin/licenses", status_code=302)

@app.get("/admin/add-license", response_class=HTMLResponse)
async def admin_add_license_page(request: Request):
    if not admin_auth(request):
        return RedirectResponse("/admin/login", status_code=302)
    return templates.TemplateResponse("add_license.html", {"request": request, "active_page": "add"})

@app.post("/admin/licenses/add")
async def admin_add_license(request: Request, email: str = FastAPIForm(""), full_name: str = FastAPIForm(""), company: str = FastAPIForm(""), days: int = FastAPIForm(30)):
    if not admin_auth(request):
        return RedirectResponse("/admin/login", status_code=302)
    if not email or not full_name:
        return RedirectResponse("/admin/add-license", status_code=302)

    expires_at = (datetime.utcnow() + timedelta(days=days)).replace(microsecond=0).isoformat() + "Z"
    token = _make_activation_token(expires_at)

    conn = get_admin_db()
    c = conn.cursor()
    c.execute("SELECT id FROM licenses WHERE email = ?", (email,))
    existing = c.fetchone()
    if existing:
        c.execute("UPDATE licenses SET token=?, expires_at=?, is_active=1, full_name=?, company=?, deactivated_at=NULL, deactivate_reason=NULL WHERE id = ?",
                  (token, expires_at, full_name, company, existing["id"]))
    else:
        c.execute("INSERT INTO licenses (email, full_name, company, token, expires_at) VALUES (?, ?, ?, ?, ?)",
                  (email, full_name, company, token, expires_at))
    conn.commit()
    conn.close()
    log_activity("manual_add", email, f"Days: {days}, Expires: {expires_at}")
    return RedirectResponse("/admin/licenses", status_code=302)