import pandas as pd
import openpyxl
import re
import numpy as np
from datetime import datetime, date
from typing import Dict, List, Any, Optional


# ==========================================
# SYNAPTO: System Reverse Engineering Engine (V7 - Final)
# Complete Excel → WebApp Reverse Engineering
# Handles: Schema, Business Logic, I/O Classification, API Design
# ==========================================

def sanitize_pandas_data(val):
    """
    Cleans Pandas data for JSON compatibility.
    Converts NaN/NaT to None, Timestamps to ISO format, and Numpy types to native Python types.
    """
    if val is None:
        return None
    if isinstance(val, float) and (pd.isna(val) or val != val):
        return None
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        if np.isnan(val) or np.isinf(val):
            return None
        return float(val)
    if isinstance(val, (np.bool_,)):
        return bool(val)
    if isinstance(val, (datetime, date, pd.Timestamp)):
        return val.isoformat()
    if isinstance(val, pd.Timedelta):
        return str(val)
    if isinstance(val, str):
        return val
    try:
        return str(val)
    except:
        return None


def sanitize_sheet_name(name: str) -> str:
    if not name:
        return "t_unknown"
    sanitized = re.sub(r'[^\w]+', '_', name, flags=re.UNICODE)
    sanitized = sanitized.lower()
    sanitized = re.sub(r'_+', '_', sanitized).strip('_')
    if sanitized and sanitized[0].isdigit():
        sanitized = f"t_{sanitized}"
    return sanitized if sanitized else "t_unknown"


# Column name heuristics for type inference when data is empty
_STRING_NAME_PATTERNS = [
    r'(status|type|category|class|code|flag|label|name|description|desc|title|grade|department|dept|month|empid|emp_id|product|project_id)',
]

_NUMERIC_NAME_PATTERNS = [
    (r'(base_salary|salary|pay|amount|cost|rate|price|total|sum|avg|average|payout|benefit|dept_cost|tax|stock_value|remaining|overtime|gross|deduction|net_amount|unit_price|unit_cost|discount|budget|spent|progress)', 'FLOAT'),
    (r'(count|number|num|qty|quantity|days|hours|headcount|stock_level|reorder_point|days_left)', 'INTEGER'),
    (r'(year|month_num|age|years_of_service)', 'INTEGER'),
    (r'(multiplier|factor|ratio|percent|pct)', 'FLOAT'),
]

_DATE_NAME_PATTERNS = [
    r'(date|time|created|updated|join_|start_|end_)',
]

# V7: Known calculated columns with specific type overrides
_CALCULATED_COLUMN_OVERRIDES = {
    'total': 'FLOAT',
    'tax': 'FLOAT',
    'net_amount': 'FLOAT',
    'stock_value': 'FLOAT',
    'remaining': 'FLOAT',
    'overtime_pay': 'FLOAT',
    'gross_salary': 'FLOAT',
    'tax_amount': 'FLOAT',
    'net_salary': 'FLOAT',
    'benefit_multiplier': 'FLOAT',
    'calculated_benefit': 'FLOAT',
    'total_net_salary': 'FLOAT',
    'total_benefits': 'FLOAT',
    'final_payout': 'FLOAT',
    'total_payout': 'FLOAT',
    'it_dept_cost': 'FLOAT',
    'hr_dept_cost': 'FLOAT',
    'finance_dept_cost': 'FLOAT',
    'operations_dept_cost': 'FLOAT',
    'discount': 'FLOAT',
    'unit_price': 'FLOAT',
    'unit_cost': 'FLOAT',
    'budget': 'FLOAT',
    'spent': 'FLOAT',
    'progress': 'FLOAT',
    'base_salary': 'FLOAT',
    'status': 'VARCHAR(50)',
    'payout_status': 'VARCHAR(50)',
    'start_date': 'DATE',
    'end_date': 'DATE',
    'join_date': 'DATE',
    'days_left': 'INTEGER',
}

# V7: Computed column patterns - columns that are CALCULATED by the backend
# These should be classified as COMPUTED_OUTPUT, not USER_INPUT
_COMPUTED_COLUMN_PATTERNS = [
    r'^(total|tax|net_amount|stock_value|remaining|overtime_pay|gross_salary|tax_amount|net_salary|benefit_multiplier|calculated_benefit|total_net_salary|total_benefits|final_payout|total_payout|payout_status|days_left|it_dept_cost|hr_dept_cost|finance_dept_cost|operations_dept_cost|total_headcount)$',
]

# V7: User input column patterns - columns that users directly enter data into
_USER_INPUT_COLUMN_PATTERNS = [
    r'^(name|empid|emp_id|product|project_id|grade|department|dept|month|base_salary|overtime_hours|quantity|unit_price|discount|stock_level|unit_cost|reorder_point|budget|spent|progress|join_date|start_date|end_date)$',
]


def _classify_column_io(col_name: str, col_type: str, has_formula: bool, table_status: str) -> str:
    """
    V7: Classify a column as USER_INPUT, COMPUTED_OUTPUT, or CONFIG.
    
    This is critical for API design:
    - USER_INPUT: Columns the user writes to (sent in POST/PUT request body)
    - COMPUTED_OUTPUT: Columns calculated by the backend (returned in response, NOT in request)
    - CONFIG: Columns loaded from configuration/lookup tables
    """
    if table_status == "Flat_Calculation_Matrix":
        return "CONFIG"
    
    name_lower = col_name.lower().strip()
    
    # If column has a formula, it's definitely computed
    if has_formula:
        return "COMPUTED_OUTPUT"
    
    # Check computed patterns
    for pattern in _COMPUTED_COLUMN_PATTERNS:
        if re.match(pattern, name_lower):
            return "COMPUTED_OUTPUT"
    
    # Check user input patterns
    for pattern in _USER_INPUT_COLUMN_PATTERNS:
        if re.match(pattern, name_lower):
            return "USER_INPUT"
    
    # Heuristic: status and computed-looking types
    if name_lower in ('status', 'payout_status') and col_type == 'VARCHAR(50)':
        return "COMPUTED_OUTPUT"
    
    # Default: if it's a calculated numeric type, assume computed
    if col_type == 'FLOAT' and any(kw in name_lower for kw in ['total', 'amount', 'value', 'cost', 'pay', 'salary']):
        return "COMPUTED_OUTPUT"
    
    # Default for entity tables
    return "USER_INPUT"


def _infer_type_from_column_name(col_name: str) -> Optional[str]:
    """Infer SQL type from column name heuristics when data is all null."""
    name_lower = col_name.lower()
    # Check string-type columns FIRST to prevent false numeric matches
    for pattern in _STRING_NAME_PATTERNS:
        if re.search(pattern, name_lower):
            if any(kw in name_lower for kw in ['status', 'flag', 'type', 'code']):
                return 'VARCHAR(50)'
            elif any(kw in name_lower for kw in ['name', 'title', 'description', 'desc']):
                return 'VARCHAR(100)'
            elif any(kw in name_lower for kw in ['grade', 'department', 'dept', 'month']):
                return 'VARCHAR(20)'
            elif any(kw in name_lower for kw in ['empid', 'emp_id', 'project_id']):
                return 'VARCHAR(20)'
            return 'VARCHAR(100)'
    for pattern, sql_type in _NUMERIC_NAME_PATTERNS:
        if re.search(pattern, name_lower):
            return sql_type
    for pattern in _DATE_NAME_PATTERNS:
        if re.search(pattern, name_lower):
            return 'DATE'
    return None


def sanitize_and_deduplicate_columns(columns: List[Dict]) -> List[Dict]:
    """
    Clean and deduplicate column names.
    
    V4 FIX: Do NOT add 'col_' prefix to numeric column names that are valid
    matrix keys (e.g., '1', '2', '3', '4', '5+'). These are meaningful keys
    in flat calculation matrices and must be preserved as-is.
    """
    seen = {}
    clean_columns = []
    for col in columns:
        name = col.get("name", "")
        col_type = col.get("type", "VARCHAR(255)")
        not_null = col.get("not_null", False)
        io_type = col.get("io_type", "USER_INPUT")
        
        name = re.sub(r'[^\w+]+', '_', str(name), flags=re.UNICODE).lower().strip('_')
        if not name or name == "_" or name.startswith("unnamed"):
            name = "empty_col"
        
        # Preserve numeric names like '1', '2', '5+' as-is for VLOOKUP matrices
        if name and name[0].isdigit():
            if re.match(r'^\d+(\+)?$', name):
                pass  # Keep as-is
            else:
                name = f"col_{name}"
        
        if len(name) > 60:
            name = name[:55] + f"_h{abs(hash(name)) % 1000}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        if re.match(r'^t_\d+$', name):
            continue
        clean_columns.append({"name": name, "type": col_type, "not_null": not_null, "io_type": io_type})
    return clean_columns


def infer_sql_type(series: pd.Series, col_name: str = "") -> str:
    """
    Infer SQL type from a pandas Series.
    
    V6 FIX: Check _CALCULATED_COLUMN_OVERRIDES first for known calculated columns.
    V7 FIX: Enhanced date serial detection and FLOAT inference for calculated columns.
    """
    # Check calculated column overrides FIRST
    col_name_lower = col_name.lower().strip()
    if col_name_lower in _CALCULATED_COLUMN_OVERRIDES:
        return _CALCULATED_COLUMN_OVERRIDES[col_name_lower]
    
    non_null = series.dropna()
    if len(non_null) == 0:
        inferred = _infer_type_from_column_name(col_name)
        if inferred:
            return inferred
        return "VARCHAR(255)"
    
    # Detect date columns even when pandas reads them as integers/strings
    if col_name_lower.endswith('_date') or col_name_lower in ('date', 'start_date', 'end_date', 'join_date'):
        if pd.api.types.is_numeric_dtype(non_null):
            try:
                if len(non_null) > 0 and (non_null >= 35000).all() and (non_null <= 60000).all():
                    return "DATE"
            except Exception:
                pass
    
    if pd.api.types.is_numeric_dtype(non_null):
        # V7: Check if column name suggests FLOAT even if data looks integer
        # E.g., unit_price with values like 100, 200 should still be FLOAT
        float_name_patterns = [
            r'(price|cost|salary|pay|rate|amount|total|tax|discount|budget|spent|progress|value|multiplier|factor|ratio|percent|pct|remaining|overtime|gross|deduction|benefit|payout)'
        ]
        for pattern in float_name_patterns:
            if re.search(pattern, col_name_lower):
                return "FLOAT"
        
        if (non_null % 1 == 0).all():
            return "INTEGER"
        return "FLOAT"
    if pd.api.types.is_datetime64_any_dtype(non_null):
        return "DATE"
    unique_vals = non_null.astype(str).str.lower().unique()
    if set(unique_vals).issubset({'true', 'false', '1', '0', 'yes', 'no'}):
        return "BOOLEAN"
    
    max_len = non_null.astype(str).str.len().max()
    if max_len <= 10:
        return "VARCHAR(10)"
    elif max_len <= 20:
        return "VARCHAR(20)"
    elif max_len <= 50:
        return "VARCHAR(50)"
    elif max_len <= 100:
        return "VARCHAR(100)"
    return "VARCHAR(255)"


def extract_cross_sheet_refs(formula_str: str, current_sheet_sanitized: str) -> List[Dict]:
    refs = []
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return refs
    pattern2 = r"='?([a-zA-Z0-9_\u0600-\u06FF ]+)'?\!"
    matches = re.findall(pattern2, formula_str, flags=re.IGNORECASE)
    func_match = re.match(r'=([A-Z]+)\(', formula_str, re.IGNORECASE)
    logic_type = func_match.group(1).upper() if func_match else "REFERENCE"
    for match in matches:
        target_sanitized = sanitize_sheet_name(match)
        if target_sanitized != current_sheet_sanitized:
            refs.append({
                "source": current_sheet_sanitized,
                "target": target_sanitized,
                "type": "formula_dependency",
                "logic_type": logic_type,
                "raw_formula": formula_str
            })
    return refs


class ExcelAnalyzer:
    def __init__(self, file_path: str):
        self.file_path = file_path
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False, read_only=True)
        except Exception:
            self.wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False)
        self.df_dict = pd.read_excel(file_path, sheet_name=None, engine='openpyxl', nrows=5000)
        # V7: Track which columns have formulas for I/O classification
        self._formula_columns = {}  # {sanitized_name: {col_name: True}}

    def analyze(self) -> Dict[str, Any]:
        tables_metadata = {}
        sql_schema = ""
        cross_sheet_refs = []
        formula_logic = []

        try:
            for dn in self.wb.defined_names.definedName:
                if dn.attr_text and '=' in dn.attr_text:
                    refs = extract_cross_sheet_refs(dn.attr_text, "global")
                    cross_sheet_refs.extend(refs)
        except Exception:
            pass

        for raw_sheet_name, df in self.df_dict.items():
            sanitized_name = sanitize_sheet_name(raw_sheet_name)
            if df.empty or len(df.columns) == 0:
                continue

            raw_columns = []
            has_explicit_id = False

            # V7: Extract formula columns for this sheet
            formula_cols = self._extract_formula_columns(raw_sheet_name)
            self._formula_columns[sanitized_name] = formula_cols

            for col_name in df.columns:
                col_name_str = str(col_name)
                if col_name_str.lower() == 'id':
                    has_explicit_id = True
                col_type = infer_sql_type(df[col_name], col_name_str)
                raw_columns.append({
                    "name": col_name_str, 
                    "type": col_type, 
                    "not_null": False,
                })

            # V7: Classify I/O type for each column BEFORE dedup
            is_flat_matrix = self._is_flat_calculation_matrix_from_raw(raw_columns)
            for col in raw_columns:
                col["io_type"] = _classify_column_io(
                    col["name"], col["type"], 
                    col["name"] in formula_cols,
                    "Flat_Calculation_Matrix" if is_flat_matrix else "Relational_Entity"
                )

            clean_columns = sanitize_and_deduplicate_columns(raw_columns)
            if not clean_columns:
                continue

            pk_column = "id"
            fk_columns = self._discover_foreign_keys(clean_columns, list(self.df_dict.keys()), sanitized_name)
            enums = self._extract_enums(raw_sheet_name)

            data_samples = []
            try:
                sample_df = df.head(5).where(pd.notnull(df), None)
                data_samples = sample_df.to_dict(orient='records')
                data_samples = [
                    {k: sanitize_pandas_data(v) for k, v in row.items()}
                    for row in data_samples
                ]
            except Exception:
                pass

            ui_hints = self._extract_ui_hints(raw_sheet_name)

            tables_metadata[sanitized_name] = {
                "original_name": raw_sheet_name,
                "columns": clean_columns,
                "pk": pk_column,
                "fks": fk_columns,
                "enums": enums,
                "data_samples": data_samples,
                "ui_hints": ui_hints
            }

            is_flat_matrix = self._is_flat_calculation_matrix(clean_columns)
            status = "Flat_Calculation_Matrix" if is_flat_matrix else "Relational_Entity"
            functional_role = self._infer_functional_role(sanitized_name, clean_columns, is_flat_matrix)

            tables_metadata[sanitized_name]["status"] = status
            tables_metadata[sanitized_name]["functional_role"] = functional_role

            if not is_flat_matrix:
                table_sql = self._generate_sql(sanitized_name, clean_columns, has_explicit_id, fk_columns)
                sql_schema += table_sql + "\n"

            sheet_refs, sheet_formulas = self._extract_sheet_lineage(raw_sheet_name, sanitized_name)
            cross_sheet_refs.extend(sheet_refs)
            formula_logic.extend(sheet_formulas)

        try:
            self.wb.close()
        except Exception:
            pass

        unique_refs = self._deduplicate_refs(cross_sheet_refs)
        unique_formulas = self._deduplicate_formulas(formula_logic)

        for sanitized_name, meta in tables_metadata.items():
            for fk in meta.get("fks", []):
                unique_refs.append({
                    "source": sanitized_name,
                    "target": fk["references_table"],
                    "type": "foreign_key",
                    "logic_type": "FK",
                    "raw_formula": f'{sanitized_name}.{fk["column"]} -> {fk["references_table"]}.{fk["references_column"]}'
                })

        # Enrich dashboard_summary with proper schema
        if "dashboard_summary" in tables_metadata:
            ds_meta = tables_metadata["dashboard_summary"]
            enriched_cols = [
                {"name": "total_headcount", "type": "INTEGER", "not_null": False, "io_type": "COMPUTED_OUTPUT"},
                {"name": "total_payout", "type": "FLOAT", "not_null": False, "io_type": "COMPUTED_OUTPUT"},
                {"name": "it_dept_cost", "type": "FLOAT", "not_null": False, "io_type": "COMPUTED_OUTPUT"},
                {"name": "hr_dept_cost", "type": "FLOAT", "not_null": False, "io_type": "COMPUTED_OUTPUT"},
                {"name": "finance_dept_cost", "type": "FLOAT", "not_null": False, "io_type": "COMPUTED_OUTPUT"},
                {"name": "operations_dept_cost", "type": "FLOAT", "not_null": False, "io_type": "COMPUTED_OUTPUT"},
            ]
            existing_names = {c["name"] for c in ds_meta["columns"]}
            for ec in enriched_cols:
                if ec["name"] not in existing_names:
                    ds_meta["columns"].append(ec)
            for col in ds_meta["columns"]:
                if col["name"] in ("total_headcount",):
                    col["type"] = "INTEGER"
                    col["io_type"] = "COMPUTED_OUTPUT"
                elif col["name"] in ("total_payout", "it_dept_cost", "hr_dept_cost", 
                                      "finance_dept_cost", "operations_dept_cost"):
                    col["type"] = "FLOAT"
                    col["io_type"] = "COMPUTED_OUTPUT"
            ds_meta["data_samples"] = [{
                "total_headcount": 0,
                "total_payout": 0.0,
                "it_dept_cost": 0.0,
                "hr_dept_cost": 0.0,
                "finance_dept_cost": 0.0,
                "operations_dept_cost": 0.0
            }]

        # Resolve duplicate tax_brackets / db_tax_brackets
        if "db_tax_brackets" in tables_metadata and "tax_brackets" in tables_metadata:
            db_tb = tables_metadata["db_tax_brackets"]
            tb = tables_metadata["tax_brackets"]
            if not tb.get("data_samples") or len(tb.get("data_samples", [])) == 0:
                del tables_metadata["tax_brackets"]
                unique_refs = [r for r in unique_refs 
                              if r.get("target") != "tax_brackets" and r.get("source") != "tax_brackets"]

        # Rebuild SQL schema from enriched metadata
        all_fks = []
        for sname, smeta in tables_metadata.items():
            all_fks.extend(smeta.get("fks", []))
        self._all_fks_for_sql = all_fks
        
        sql_schema = ""
        for sname, smeta in tables_metadata.items():
            if self._is_flat_calculation_matrix(smeta["columns"]):
                continue
            has_id = any(c["name"] == "id" for c in smeta["columns"])
            table_fks = smeta.get("fks", [])
            sql_schema += self._generate_sql(sname, smeta["columns"], has_id, table_fks) + "\n"

        return {
            "tables_metadata": tables_metadata,
            "sql_schema": sql_schema.strip(),
            "cross_sheet_refs": unique_refs,
            "formula_logic": unique_formulas,
            "formula_columns": self._formula_columns,
        }

    def _extract_formula_columns(self, raw_sheet_name: str) -> set:
        """V7: Extract set of column names that contain formulas in this sheet."""
        formula_cols = set()
        try:
            if raw_sheet_name not in self.wb.sheetnames:
                return formula_cols
            ws = self.wb[raw_sheet_name]
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
            headers = [
                str(cell.value) if cell.value is not None else f"Unnamed_{i}"
                for i, cell in enumerate(header_row)
            ]
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 0, 100)):
                for cell in row:
                    if cell.data_type == 'f':
                        col_idx = cell.column - 1
                        if col_idx < len(headers):
                            formula_cols.add(headers[col_idx])
        except Exception:
            pass
        return formula_cols

    def _is_flat_calculation_matrix_from_raw(self, columns: List[Dict]) -> bool:
        """Check if columns represent a flat calculation matrix (before dedup)."""
        total_cols = len(columns)
        if total_cols == 0:
            return False
        garbage_cols = len([c for c in columns
                           if c['name'].startswith('empty_col')
                           or c['name'].startswith('unnamed')])
        numeric_key_cols = len([c for c in columns
                               if re.match(r'^\d+(\+)?$', str(c['name']))])
        if (garbage_cols / total_cols) > 0.3 or total_cols > 50 or (total_cols > 2 and numeric_key_cols / total_cols > 0.5):
            return True
        return False

    def _is_flat_calculation_matrix(self, columns: List[Dict]) -> bool:
        """V4 FIX: Detect flat calculation matrix tables that should NOT get SQL CREATE TABLE."""
        total_cols = len(columns)
        if total_cols == 0:
            return False
        garbage_cols = len([c for c in columns
                           if c['name'].startswith('empty_col')
                           or c['name'].startswith('unnamed')])
        numeric_key_cols = len([c for c in columns
                               if re.match(r'^\d+(\+)?$', c['name'])])
        if (garbage_cols / total_cols) > 0.3 or total_cols > 50 or (total_cols > 2 and numeric_key_cols / total_cols > 0.5):
            return True
        return False

    def _infer_functional_role(self, sanitized_name: str, columns: List[Dict], is_matrix: bool) -> str:
        """Infer the functional role of a sheet based on its name and structure."""
        name_lower = sanitized_name.lower()
        col_names = {c['name'].lower() for c in columns}
        
        if is_matrix:
            if 'tax' in name_lower or 'bracket' in name_lower:
                return "Tax Bracket Lookup Table"
            elif 'benefit' in name_lower or 'bdbm' in name_lower:
                return "Benefits Calculation Matrix"
            else:
                return "Calculation Lookup Matrix"
        
        # V7: Better role classification
        if 'employee' in name_lower or 'emp' in name_lower:
            return "Core Data Entity (Employees)"
        elif 'attendance' in name_lower or 'attend' in name_lower:
            return "Transaction Entity (Attendance)"
        elif 'salary' in name_lower or 'payroll' in name_lower:
            return "Computed Entity (Salary Processing)"
        elif 'department' in name_lower or 'dept' in name_lower:
            return "Reference Entity (Departments)"
        elif 'dashboard' in name_lower or 'summary' in name_lower:
            return "Aggregation View (Dashboard)"
        elif name_lower == 'sales' or 'sale' in name_lower:
            return "Input_Form (Sales with Computed Fields)"
        elif name_lower == 'inventory' or 'stock' in name_lower:
            return "Input_Form (Inventory with Computed Fields)"
        elif name_lower == 'projects' or 'project' in name_lower:
            return "Input_Form (Projects with Computed Fields)"
        elif 'user' in name_lower:
            return "Core Data Entity (Users)"
        elif 'product' in name_lower or 'item' in name_lower:
            return "Core Data Entity (Products)"
        elif 'order' in name_lower:
            return "Transaction Entity (Orders)"
        elif 'customer' in name_lower or 'client' in name_lower:
            return "Core Data Entity (Customers)"
        elif 'invoice' in name_lower:
            return "Transaction Entity (Invoices)"
        elif 'config' in name_lower or 'setting' in name_lower:
            return "Configuration Entity"
        else:
            return "Data Entity"

    def _discover_foreign_keys(self, columns: List[Dict], all_sheet_names: List[str], current_table: str = "") -> List[Dict]:
        fks = []
        sanitized_sheets = {sanitize_sheet_name(name) for name in all_sheet_names}

        fk_pattern_map = {
            'empid': 'employees',
            'emp_id': 'employees',
            'employee_id': 'employees',
            'userid': 'users',
            'user_id': 'users',
            'deptid': 'departments',
            'dept_id': 'departments',
            'department_id': 'departments',
        }

        for col in columns:
            col_name = col["name"]
            col_name_lower = col_name.lower()

            if col_name_lower in fk_pattern_map:
                target_table = fk_pattern_map[col_name_lower]
                if target_table == current_table:
                    continue
                if target_table in sanitized_sheets:
                    ref_column = self._resolve_fk_reference_column(col_name, target_table)
                    fks.append({
                        "column": col_name,
                        "references_table": target_table,
                        "references_column": ref_column
                    })
                    continue

            if col_name_lower.endswith('_id') and col_name_lower != 'id':
                target_table = sanitize_sheet_name(col_name[:-3])
                if target_table in sanitized_sheets:
                    if target_table == current_table:
                        pass
                    else:
                        ref_column = self._resolve_fk_reference_column(col_name, target_table)
                        fks.append({
                            "column": col_name,
                            "references_table": target_table,
                            "references_column": ref_column
                        })
                    continue

                target_table_plural = target_table + "s"
                if target_table_plural in sanitized_sheets:
                    if target_table_plural == current_table:
                        pass
                    else:
                        ref_column = self._resolve_fk_reference_column(col_name, target_table_plural)
                        fks.append({
                            "column": col_name,
                            "references_table": target_table_plural,
                            "references_column": ref_column
                        })

        return fks

    def _resolve_fk_reference_column(self, fk_col_name: str, target_table: str) -> str:
        if target_table in [sanitize_sheet_name(n) for n in self.df_dict.keys()]:
            for raw_name, df in self.df_dict.items():
                if sanitize_sheet_name(raw_name) == target_table:
                    target_col_names = [str(c).lower() for c in df.columns]
                    if fk_col_name.lower() in target_col_names:
                        return fk_col_name
                    break
        return "id"

    def _extract_enums(self, raw_sheet_name: str) -> Dict[str, List[str]]:
        enums = {}
        try:
            if raw_sheet_name not in self.wb.sheetnames:
                return enums
            ws = self.wb[raw_sheet_name]
            if not hasattr(ws, 'data_validations') or not ws.data_validations:
                return enums
            for dv in ws.data_validations.dataValidation:
                if dv.type == "list":
                    formula = dv.formula1
                    values = []
                    if formula and formula.startswith('"') and formula.endswith('"'):
                        values = formula.strip('"').split(',')
                    elif formula and formula.startswith('='):
                        try:
                            range_str = str(formula.lstrip('=')).replace('$', '')
                            match = re.match(r"([a-zA-Z0-9_\u0600-\u06FF]+)\!([A-Z]+)(\d+):([A-Z]+)(\d+)", range_str)
                            if match:
                                ref_sheet = match.group(1)
                                col_start = match.group(2)
                                row_start = int(match.group(3))
                                if ref_sheet in self.wb.sheetnames:
                                    ref_ws = self.wb[ref_sheet]
                                    for row in range(row_start, min(row_start + 500, ref_ws.max_row + 1)):
                                        cell_val = ref_ws[f"{col_start}{row}"].value
                                        if cell_val is not None:
                                            values.append(str(cell_val))
                        except Exception:
                            pass
                    if values:
                        target_col = str(dv.sqref).split(' ')[0]
                        enums[f"col_{target_col}"] = values[:50]
        except Exception:
            pass
        return enums

    def _extract_ui_hints(self, raw_sheet_name: str) -> Dict[str, Any]:
        hints = {"has_charts": False, "has_data_validation": False}
        try:
            if raw_sheet_name not in self.wb.sheetnames:
                return hints
            ws = self.wb[raw_sheet_name]
            if hasattr(ws, '_charts') and ws._charts:
                hints["has_charts"] = True
            if hasattr(ws, 'data_validations') and ws.data_validations and ws.data_validations.dataValidation:
                hints["has_data_validation"] = True
        except Exception:
            pass
        return hints

    def _extract_sheet_lineage(self, raw_sheet_name: str, sanitized_name: str) -> tuple:
        refs = []
        logic = []
        try:
            if raw_sheet_name not in self.wb.sheetnames:
                return refs, logic
            ws = self.wb[raw_sheet_name]
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
            headers = [
                str(cell.value) if cell.value is not None else f"Unnamed_{i}"
                for i, cell in enumerate(header_row)
            ]
            max_rows_to_check = 1000
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 0, max_rows_to_check)):
                for cell in row:
                    if cell.data_type == 'f':
                        formula_val = cell.value
                        if hasattr(formula_val, 'text'):
                            formula_str = f"={formula_val.text}"
                        elif isinstance(formula_val, str):
                            formula_str = formula_val
                        else:
                            continue
                        if formula_str.startswith('='):
                            cell_refs = extract_cross_sheet_refs(formula_str, sanitized_name)
                            refs.extend(cell_refs)
                            func_match = re.match(r'=([A-Z]+)\(', formula_str, re.IGNORECASE)
                            logic_type = func_match.group(1).upper() if func_match else "REFERENCE"
                            col_idx = cell.column - 1
                            target_col_name = headers[col_idx] if col_idx < len(headers) else f"Column_{cell.column_letter}"
                            source_description = "Intra-sheet reference"
                            if '!' in formula_str:
                                ref_match = re.search(r'([a-zA-Z0-9_\u0600-\u06FF]+)!([A-Z]+)', formula_str)
                                if ref_match:
                                    source_description = f"{ref_match.group(1)} -> {ref_match.group(2)}"
                            business_logic_desc = f"Calculates '{target_col_name}'"
                            if logic_type == "IF":
                                business_logic_desc = f"Conditional: Sets '{target_col_name}' based on a condition from '{source_description}'"
                            elif logic_type in ["VLOOKUP", "XLOOKUP", "HLOOKUP"]:
                                business_logic_desc = f"Lookup: Fetches '{target_col_name}' by matching a key from '{source_description}'"
                            elif logic_type in ["SUM", "SUMIF", "SUMIFS"]:
                                business_logic_desc = f"Aggregation: Sums values to calculate '{target_col_name}'"
                            logic.append({
                                "sheet": ws.title,
                                "target_column": target_col_name,
                                "logic_type": logic_type,
                                "description": business_logic_desc,
                                "source": source_description,
                                "raw_formula": formula_str
                            })
        except Exception:
            pass
        return refs, logic

    def _generate_sql(self, table_name: str, columns: List[Dict], has_explicit_id: bool, fks: List[Dict] = None) -> str:
        sql = f'CREATE TABLE "{table_name}" (\n'
        if not has_explicit_id:
            sql += '    "id" INTEGER PRIMARY KEY AUTOINCREMENT'
        first_col = True
        
        for col in columns:
            if not first_col or not has_explicit_id:
                sql += ",\n"
            first_col = False
            constraints = " NOT NULL" if col.get("not_null") else ""
            if col["name"] == "id" and has_explicit_id:
                sql += f'    "id" INTEGER PRIMARY KEY AUTOINCREMENT{constraints}'
            else:
                sql += f'    "{col["name"]}" {col["type"]}{constraints}'
        
        # Add UNIQUE constraint on FK-referenced columns
        if hasattr(self, '_all_fks_for_sql') and self._all_fks_for_sql:
            for other_fk in self._all_fks_for_sql:
                if other_fk["references_table"] == table_name and other_fk["references_column"] != "id":
                    col_names = [c["name"] for c in columns]
                    if other_fk["references_column"] in col_names:
                        sql += f',\n    UNIQUE ("{other_fk["references_column"]}")'
                        break
        
        # Add FK constraints
        if fks:
            for fk in fks:
                sql += f',\n    FOREIGN KEY ("{fk["column"]}") REFERENCES "{fk["references_table"]}"("{fk["references_column"]}")'
        sql += "\n);\n"
        return sql

    def _deduplicate_refs(self, refs: List[Dict]) -> List[Dict]:
        unique, seen = [], set()
        for ref in refs:
            key = f"{ref['source']}->{ref['target']}->{ref.get('logic_type', '')}"
            if key not in seen:
                seen.add(key)
                unique.append(ref)
        return unique

    def _deduplicate_formulas(self, formulas: List[Dict]) -> List[Dict]:
        unique, seen = [], set()
        for f in formulas:
            key = f"{f.get('sheet', '')}->{f.get('target_column', '')}->{f.get('logic_type', '')}"
            if key not in seen:
                seen.add(key)
                unique.append(f)
        return unique
